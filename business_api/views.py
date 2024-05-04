import string
import hmac
from decouple import config
from django.db import IntegrityError
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt

from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.authtoken.models import Token

from business_api import models
import hashlib

from rest_framework.authentication import TokenAuthentication

import datetime
import json
import random
from time import sleep

import requests
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from django.conf import settings

import firebase_admin
from firebase_admin import credentials
from firebase_admin import db, firestore

from business_api.models import MTNTransaction

if not firebase_admin._apps:
    cred = credentials.Certificate(settings.FIREBASE_ADMIN_CERT)
    firebase_admin.initialize_app(cred, {
        'databaseURL': 'https://bestpay-flutter-default-rtdb.firebaseio.com'
    })

database = firestore.client()
user_collection = database.collection(u'Users')
cashback_collection = database.collection(u"general_cashback")
history_collection = database.collection(u'History Web')
mail_collection = database.collection('mail')
mtn_history = database.collection('MTN_Admin_History')
mtn_tranx = mtn_history.document('mtnTransactions')
big_time = mtn_tranx.collection('big_time')
mtn_other = mtn_tranx.collection('mtnOther')
bearer_token_collection = database.collection("_KeysAndBearer")
history_web = database.collection(u'History Web').document('all_users')

totals_collection = database.collection('Totals')
admin_collection = database.collection('Admin')


class BearerTokenAuthentication(TokenAuthentication):
    keyword = 'Bearer'


def tranx_id_generator():
    file = open('business_api/counter.txt', 'r')
    content = file.read()

    tranx_id = int(content) + 1
    file = open('business_api/counter.txt', 'w')
    file.write(str(tranx_id))
    file = open('business_api/counter.txt', 'r')
    content = file.read()
    print(content)
    return content


def generate_tokenn(length):
    characters = string.ascii_letters + string.digits
    return ''.join(random.choice(characters) for _ in range(length))


def check_user_balance_against_price(user_id, price):
    details = get_user_details(user_id)
    wallet_balance = details['wallet']
    if wallet_balance is not None:
        return wallet_balance >= float(price)
    else:
        return None


def check_user_at_balance_against_price(user_id, data_volume):
    details = get_user_details(user_id)
    at_balance = details['at_balance']
    if at_balance is not None:
        return at_balance >= float(data_volume)
    else:
        return None


def get_user_details(user_id):
    user = user_collection.document(user_id)
    doc = user.get()
    if doc.exists:
        doc_dict = doc.to_dict()
        print(doc_dict)
        first_name = doc_dict['first name']
        last_name = doc_dict['last name']
        email = doc_dict['email']
        phone = doc_dict['phone']
        print(first_name), print(last_name), print(email), print(phone)
        return doc.to_dict()
    else:
        return None


def send_ishare_bundle(first_name: str, last_name: str, buyer, receiver: str, email: str, bundle: float
                       ):
    url = "https://backend.boldassure.net:445/live/api/context/business/transaction/new-transaction"

    payload = json.dumps({
        "accountNo": buyer,
        "accountFirstName": first_name,
        "accountLastName": last_name,
        "accountMsisdn": receiver,
        "accountEmail": email,
        "accountVoiceBalance": 0,
        "accountDataBalance": bundle,
        "accountCashBalance": 0,
        "active": True
    })

    token = bearer_token_collection.document("Active_API_BoldAssure")
    token_doc = token.get()
    token_doc_dict = token_doc.to_dict()
    tokennn = token_doc_dict['ishare_bearer']
    print(tokennn)

    headers = {
        'Authorization': tokennn,
        'Content-Type': 'application/json'
    }

    response = requests.request("POST", url, headers=headers, data=payload)
    print(
        f"{response.json()} from first method +++++++++++++++++++++++++++++++++++++==========================================")
    return response


def ishare_verification(batch_id):
    if batch_id == "No batchId":
        return False

    url = f"https://backend.boldassure.net:445/live/api/context/business/airteltigo-gh/ishare/tranx-status/{batch_id}"

    payload = {}
    token = bearer_token_collection.document("Active_API_BoldAssure")
    token_doc = token.get()
    token_doc_dict = token_doc.to_dict()
    tokennn = token_doc_dict['ishare_bearer']
    headers = {
        'Authorization': tokennn
    }

    response = requests.request("GET", url, headers=headers, data=payload)

    if response.status_code == 200:
        json_data = response.json()
        print(json_data)
        return json_data
    else:
        return False


def send_and_save_to_history(user_id,
                             data_volume: float, reference: str, amount: float, receiver: str,
                             date: str, time: str, date_and_time: str):
    user_details = get_user_details(user_id)
    first_name = user_details['first name']
    last_name = user_details['last name']
    email = user_details['email']
    phone = user_details['phone']
    wallet = user_details['wallet']

    data = {
        'batch_id': "unknown",
        'buyer': phone,
        'color_code': "Green",
        'amount': amount,
        'data_break_down': data_volume,
        'data_volume': data_volume,
        'date': date,
        'date_and_time': date_and_time,
        'done': "unknown",
        'email': email,
        'image': user_id,
        'ishareBalance': 0,
        'name': f"{first_name} {last_name}",
        'number': receiver,
        'paid_at': date_and_time,
        'reference': reference,
        'responseCode': "0",
        'status': "Delivered",
        'time': time,
        'tranxId': str(tranx_id_generator()),
        'type': "AT PREMIUM BUNDLE",
        'uid': user_id,
        'bal': wallet
    }
    history_collection.document(date_and_time).set(data)
    history_web.collection(email).document(date_and_time).set(data)

    print("first save")

    ishare_response = send_ishare_bundle(first_name=first_name, last_name=last_name, receiver=receiver,
                                         buyer=phone,
                                         bundle=data_volume,
                                         email=email)
    json_response = ishare_response.json()
    print(f"hello:{json_response}")
    print(ishare_response.status_code)
    try:
        batch_id = json_response["batchId"]
    except KeyError:
        batch_id = None
    print(batch_id)

    doc_ref = history_collection.document(date_and_time)
    doc_ref.update({'batch_id': batch_id, 'responseCode': ishare_response.status_code})
    history_web.collection(email).document(date_and_time).update(
        {'batch_id': batch_id, 'responseCode': ishare_response.status_code})
    # data = {
    #     'batch_id': batch_id,
    #     'buyer': phone,
    #     'color_code': color_code,
    #     'amount': amount,
    #     'data_break_down': data_break_down,
    #     'data_volume': data_volume,
    #     'date': date,
    #     'date_and_time': date_and_time,
    #     'done': "unknown",
    #     'email': email,
    #     'image': image,
    #     'ishareBalance': ishare_balance,
    #     'name': f"{first_name} {last_name}",
    #     'number': receiver,
    #     'paid_at': paid_at,
    #     'reference': reference,
    #     'responseCode': status_code,
    #     'status': txn_status,
    #     'time': time,
    #     'tranxId': str(tranx_id_gen()),
    #     'type': txn_type,
    #     'uid': user_id
    # }
    # history_collection.document(date_and_time).set(data)
    # history_web.collection(email).document(date_and_time).set(data)

    print("firebase saved")
    return ishare_response


def big_time_transaction(receiver, date, time, date_and_time, phone, amount, data_volume, details: dict, ref,
                         channel, txn_status, user_id):
    print("==========")
    print(amount)
    data = {
        'batch_id': "unknown",
        'buyer': phone,
        'color_code': "Green",
        'amount': amount,
        'data_break_down': str(data_volume),
        'data_volume': data_volume,
        'date': date,
        'date_and_time': date_and_time,
        'done': "unknown",
        'email': details['email'],
        'image': details['user_id'],
        'ishareBalance': 0,
        'name': f"{details['first_name']} {details['last_name']}",
        'number': receiver,
        'paid_at': str(date_and_time),
        'reference': ref,
        'responseCode': 200,
        'status': txn_status,
        'time': time,
        'tranxId': str(tranx_id_generator()),
        'type': "AT Big Time",
        'uid': details['user_id']
    }
    history_collection.document(date_and_time).set(data)
    history_web.collection(details['email']).document(date_and_time).set(data)
    big_time.document(date_and_time).set(data)
    user = history_collection.document(date_and_time)
    doc = user.get()
    print(doc.to_dict())
    tranx_id = doc.to_dict()['tranxId']
    mail_doc_ref = mail_collection.document()
    file_path = 'business_api/mtn_maill.txt'  # Replace with your file path

    name = details['first_name']
    volume = data_volume
    date = date_and_time
    reference_t = ref
    receiver_t = receiver

    # tot = user_collection.document(user_id)
    # print(tot.get().to_dict())
    # try:
    #     print(tot.get().to_dict()['bt_total_sales'])
    #     previous_sale = tot.get().to_dict()['bt_total_sales']
    #     print(f"Previous Sale: {previous_sale}")
    #     new_sale = float(previous_sale) + float(amount)
    #     print(new_sale)
    #     user_collection.document(user_id).update({'bt_total_sales': new_sale})
    # except:
    #     user_collection.document(user_id).update({'bt_total_sales': amount})

    # tat = cashback_collection.document(user_id)
    # print(tat.get().to_dict())
    #
    # try:
    #     previous_cashback = tat.get().to_dict()['cashback_wallet']
    #     print(previous_cashback)
    #     cashback_balance = (0.5 / 100) * float(amount)
    #     print(cashback_balance)
    #     new_cashback = float(previous_cashback) + float(cashback_balance)
    #     print(new_cashback)
    #     cashback_collection.document(user_id).update(
    #         {'cashback_wallet': new_cashback, 'phone_number': phone})
    # except TypeError as e:
    #     print(e)
    #     cashback_balance = (0.5 / 100) * float(amount)
    #     print(cashback_balance)
    #     cashback_collection.document(user_id).set(
    #         {'cashback_wallet': cashback_balance, 'phone_number': phone})
    #
    #     print(cashback_collection.document(user_id).get().to_dict())
    #     print("did")

    # previous_big_time_totals = totals_collection.document('BIGTIME TOTALS')
    # all_totals = totals_collection.document('ALL TOTALS')
    # doc = previous_big_time_totals.get()
    # doc_dict = doc.to_dict()
    #
    # previous_total_trans = doc_dict['total_trans']
    # previous_total_amount = doc_dict['total_amount']
    #
    # all_total_doc = all_totals.get()
    # all_total_doc_dict = all_total_doc.to_dict()
    #
    # previous_all_total_amount = all_total_doc_dict['total_amount']
    #
    # try:
    #     new_total_amount = previous_total_amount + amount
    # except:
    #     new_total_amount = amount
    #
    # try:
    #     new_total_trans = previous_total_trans + 1
    # except:
    #     new_total_trans = 1
    #
    # data = {
    #     'total_trans': new_total_trans,
    #     'total_amount': new_total_amount
    # }
    #
    # totals_collection.document('BIGTIME TOTALS').update(data)
    # totals_collection.document('ALL TOTALS').update({'total_amount': previous_all_total_amount + amount})

    with open(file_path, 'r') as file:
        html_content = file.read()

    placeholders = {
        '{name}': name,
        '{volume}': volume,
        '{date}': date,
        '{reference}': reference_t,
        '{receiver}': receiver_t
    }

    for placeholder, value in placeholders.items():
        html_content = html_content.replace(placeholder, str(value))

    mail_doc_ref.set({
        'to': details['email'],
        'message': {
            'subject': 'Big Time Data',
            'html': html_content,
            'messageId': 'CloudHub GH'
        }
    })
    return Response(data={'code': '0000', 'message': "Transaction Saved"}, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@authentication_classes([BearerTokenAuthentication])
def home(request):
    authorization_header = request.headers.get('Authorization')
    if authorization_header:
        auth_type, token = authorization_header.split(' ')
        if auth_type == 'Bearer':
            try:
                token_obj = Token.objects.get(key=token)
                user = token_obj.user
                user_details = {
                    'username': user.username,
                    'email': user.email,
                    # Add other user details as needed
                }
                return Response({'user_details': user_details, 'message': 'Welcome! You have accessed the home view.'},
                                status=status.HTTP_200_OK)
            except Token.DoesNotExist:
                return Response({'error': 'Token does not exist.'}, status=status.HTTP_401_UNAUTHORIZED)
        else:
            return Response({'error': 'Invalid Header'})
    return Response({'error': 'Authorization header missing or invalid.'}, status=status.HTTP_401_UNAUTHORIZED)


@api_view(['POST'])
@authentication_classes([])
@permission_classes([AllowAny])
def generate_token(request):
    username = request.data.get('username')
    user_id = request.data.get('user_id')
    full_name = request.data.get('full_name')
    email = request.data.get('email')
    if username and user_id and full_name and email:
        try:
            user = models.CustomUser.objects.create_user(username=username, user_id=user_id,
                                                         full_name=full_name, email=email)
            token_key = generate_tokenn(35)
            token = Token.objects.create(user=user, key=token_key)
            return Response({'token': token.key, 'message': 'Token Generation Successful'}, status=status.HTTP_200_OK)
        except IntegrityError:
            return Response({'message': 'User already exists!'}, status=status.HTTP_409_CONFLICT)
    else:
        return Response({'error': 'Username and password are required.'}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@authentication_classes([])
@permission_classes([AllowAny])
def regenerate_token(request):
    user_id = request.data.get('user_id')
    try:
        user = models.CustomUser.objects.get(user_id=user_id)
        try:
            token = Token.objects.get(user=user)
            token.delete()
        except Token.DoesNotExist:
            pass

        token_key = generate_tokenn(35)
        token = Token.objects.create(user=user, key=token_key)
        return Response({'user_id': user_id, 'token': token.key, 'message': 'Token Generation Successful'},
                        status=status.HTTP_200_OK)
    except models.CustomUser.DoesNotExist:
        return Response({'error': 'User does not exist.'}, status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@authentication_classes([BearerTokenAuthentication])
def initiate_mtn_transaction(request):
    authorization_header = request.headers.get('Authorization')
    if authorization_header:
        auth_type, token = authorization_header.split(' ')
        if auth_type == 'Bearer':
            try:
                token_obj = Token.objects.get(key=token)
                user = token_obj.user
                user_id = user.user_id
                print(user_id)

                receiver = request.data.get('receiver')
                print(receiver)
                data_volume = request.data.get('data_volume')
                print(data_volume)
                reference = request.data.get('reference')
                amount = request.data.get('amount')
                phone_number = request.data.get('phone_number')

                if not receiver or not data_volume or not reference or not amount:
                    return Response({'message': 'Body parameters not valid. Check and try again.'},
                                    status=status.HTTP_400_BAD_REQUEST)

                prices_dict = {
                    1000: 3.8,
                    2000: 7.5,
                    3000: 11.0,
                    4000: 14.3,
                    5000: 17.8,
                    6000: 20.8,
                    7000: 24.5,
                    8000: 26.0,
                    10000: 32.0,
                    15000: 47.0,
                    20000: 63.0,
                    25000: 78.0,
                    30000: 93.5,
                    40000: 127.0,
                    50000: 154.0,
                    100000: 288.0
                }

                amount_to_be_deducted = prices_dict[data_volume]
                print(str(amount_to_be_deducted) + "================")
                # channel = phone_number
                date = datetime.datetime.now().strftime("%a, %b %d, %Y")
                time = datetime.datetime.now().strftime("%I:%M:%S %p")
                date_and_time = datetime.datetime.now().isoformat()
                if "wallet" == "wallet":
                    print("used this")
                    try:
                        enough_balance = check_user_balance_against_price(user_id, amount_to_be_deducted)
                    except:
                        return Response(
                            {'code': '0001', 'message': f'User ID does not exist: User ID provided: {user_id}.'},
                            status=status.HTTP_400_BAD_REQUEST)
                else:
                    enough_balance = True
                    print("not wallet")
                print(enough_balance)
                if enough_balance:
                    user_details = get_user_details(user_id)
                    first_name = user_details['first name']
                    last_name = user_details['last name']
                    email = user_details['email']
                    phone = user_details['phone']
                    bal = user_details['wallet']
                    # hist = history_web.collection(email).document(date_and_time)
                    # doc = hist.get()
                    # if doc.exists:
                    #     print(doc)
                    #     return redirect(f"https://{callback_url}")
                    # else:
                    #     print("no record found")
                    if "wallet" == "wallet":
                        print("updated")
                        user = get_user_details(user_id)
                        if user is None:
                            return None
                        previous_user_wallet = user['wallet']
                        print(f"previous wallet: {previous_user_wallet}")
                        new_balance = float(previous_user_wallet) - float(amount_to_be_deducted)
                        print(f"new_balance:{new_balance}")
                        doc_ref = user_collection.document(user_id)
                        doc_ref.update({'wallet': new_balance})
                        user = get_user_details(user_id)
                        new_user_wallet = user['wallet']
                        print(f"new_user_wallet: {new_user_wallet}")
                        if new_user_wallet == previous_user_wallet:
                            user = get_user_details(user_id)
                            if user is None:
                                return None
                            previous_user_wallet = user['wallet']
                            print(f"previous wallet: {previous_user_wallet}")
                            new_balance = float(previous_user_wallet) - float(amount_to_be_deducted)
                            print(f"new_balance:{new_balance}")
                            doc_ref = user_collection.document(user_id)
                            doc_ref.update({'wallet': new_balance})
                            user = get_user_details(user_id)
                            new_user_wallet = user['wallet']
                            print(f"new_user_wallet: {new_user_wallet}")
                        else:
                            print("it's fine")

                    data = {
                        'batch_id': "unknown",
                        'buyer': phone_number,
                        'color_code': "Green",
                        'amount': amount_to_be_deducted,
                        'data_break_down': data_volume,
                        'data_volume': data_volume,
                        'date': str(date),
                        'date_and_time': str(date_and_time),
                        'done': "unknown",
                        'email': email,
                        'image': user_id,
                        'ishareBalance': '',
                        'name': f"{first_name} {last_name}",
                        'number': receiver,
                        'paid_at': str(date_and_time),
                        'reference': reference,
                        'responseCode': 200,
                        'status': "Undelivered",
                        'bal': bal,
                        'time': str(time),
                        'tranxId': str(tranx_id_generator()),
                        'type': "MTN Master Bundle",
                        'uid': user_id
                    }

                    history_collection.document(date_and_time).set(data)
                    history_web.collection(email).document(date_and_time).set(data)
                    user = history_collection.document(date_and_time)
                    doc = user.get()
                    print(doc.to_dict())
                    tranx_id = doc.to_dict()['tranxId']
                    second_data = {
                        'amount': amount_to_be_deducted,
                        'batch_id': "unknown",
                        'channel': "wallet",
                        'color_code': "Green",
                        'created_at': date_and_time,
                        'data_volume': data_volume,
                        'date': str(date),
                        'email': email,
                        'date_and_time': date_and_time,
                        'image': user_id,
                        'ip_address': "",
                        'ishareBalance': 0,
                        'name': f"{first_name} {last_name}",
                        'number': receiver,
                        'buyer': phone_number,
                        'paid_at': date_and_time,
                        'payment_status': "success",
                        'reference': reference,
                        'status': "Undelivered",
                        'bal': bal,
                        'time': str(time),
                        'tranxId': tranx_id,
                        'type': "MTN Master Bundle"
                    }
                    mtn_other.document(date_and_time).set(second_data)
                    print("pu")

                    # tot = user_collection.document(user_id)
                    # print(tot.get().to_dict())
                    # try:
                    #     print(tot.get().to_dict()['mtn_total_sales'])
                    #     previous_sale = tot.get().to_dict()['mtn_total_sales']
                    #     print(f"Previous Sale: {previous_sale}")
                    #     new_sale = float(previous_sale) + float(amount_to_be_deducted)
                    #     print(new_sale)
                    #     user_collection.document(user_id).update({'mtn_total_sales': new_sale})
                    # except:
                    #     user_collection.document(user_id).update({'mtn_total_sales': amount_to_be_deducted})

                    # tat = cashback_collection.document(user_id)
                    # print(tat.get().to_dict())
                    #
                    # try:
                    #     previous_cashback = tat.get().to_dict()['cashback_wallet']
                    #     print(previous_cashback)
                    #     cashback_balance = (0.5 / 100) * float(amount_to_be_deducted)
                    #     print(cashback_balance)
                    #     new_cashback = float(previous_cashback) + float(cashback_balance)
                    #     print(new_cashback)
                    #     cashback_collection.document(user_id).update(
                    #         {'cashback_wallet': new_cashback, 'phone_number': user_details['phone']})
                    #
                    # except TypeError as e:
                    #     print(e)
                    #     cashback_balance = (0.5 / 100) * float(amount_to_be_deducted)
                    #     print(cashback_balance)
                    #     cashback_collection.document(user_id).set(
                    #         {'cashback_wallet': cashback_balance, 'phone_number': user_details['phone']})
                    #
                    #     print(cashback_collection.document(user_id).get().to_dict())
                    #     print("did")

                    mail_doc_ref = mail_collection.document()
                    file_path = 'business_api/mtn_maill.txt'  # Replace with your file path

                    name = first_name
                    volume = data_volume
                    date = date_and_time
                    reference_t = reference
                    receiver_t = receiver

                    with open(file_path, 'r') as file:
                        html_content = file.read()

                    placeholders = {
                        '{name}': name,
                        '{volume}': volume,
                        '{date}': date,
                        '{reference}': reference_t,
                        '{receiver}': receiver_t
                    }

                    for placeholder, value in placeholders.items():
                        html_content = html_content.replace(placeholder, str(value))

                    mail_doc_ref.set({
                        'to': email,
                        'message': {
                            'subject': 'MTN Data',
                            'html': html_content,
                            'messageId': 'CloudHub GH'
                        }
                    })
                    print("got to redirect")
                    return Response(data={"status": "200", "message": "Transaction received successfully"},
                                    status=status.HTTP_200_OK)
                else:
                    return Response({"status": '400', 'message': 'Not enough balance to perform transaction'},
                                    status=status.HTTP_400_BAD_REQUEST)
            except Token.DoesNotExist:
                return Response({'error': 'Token does not exist.'}, status=status.HTTP_401_UNAUTHORIZED)
        else:
            return Response({'error': 'Invalid Header Provided.'}, status=status.HTTP_401_UNAUTHORIZED)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@authentication_classes([BearerTokenAuthentication])
def admin_initiate_mtn_transaction(request):
    # allowed_hosts = ['cloudhubgh.com', 'reseller.cloudhubgh.com', "api.cloudhubgh.com", "merchant.cloudhubgh.com"]
    # request_host = request.headers.get('Host')
    #
    # print(f"hosssssssssstttttttttttttttt issssssssssssss {request_host}")
    #
    # if request_host not in allowed_hosts:
    #     response1 = requests.get(
    #         f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UmpEc1JzeFV4cERKTWxUWktqZEs&to=0592117523&from=Bundle&sms=Invalid header host={request_host}")
    #     print(response1.text)
    #     return JsonResponse({'error': 'Host not allowed'}, status=403)
    allowed_origins = ['https://reseller.cloudhubgh.com']

    if 'HTTP_ORIGIN' not in request.META:
        return JsonResponse({'error': 'Origin header missing'}, status=400)

    request_origin = request.META['HTTP_ORIGIN']
    print(f"origiiiiiiiiiiiiiiinnnnnnnnnnnnnnnnn issssssssssssssssssssssssssssssssssssssssss {request_origin}")

    if request_origin not in allowed_origins:
        return JsonResponse({'error': 'Origin not allowed'}, status=403)
    authorization_header = request.headers.get('Authorization')
    if authorization_header:
        auth_type, token = authorization_header.split(' ')
        if auth_type == 'Bearer':
            try:
                token_obj = Token.objects.get(key=token)
                token_key = token_obj.key
                if token_key != config("TOKEN_KEY"):
                    return Response({'message': 'Authorisation Failed.'},
                                    status=status.HTTP_400_BAD_REQUEST)

                receiver = request.data.get('receiver')
                print(receiver)
                data_volume = request.data.get('data_volume')
                print(data_volume)
                reference = request.data.get('reference')
                user_id = request.data.get('user_id')
                amount = request.data.get('amount')
                # phone_number = request.data.get('phone_number')

                if not receiver or not data_volume or not reference or not amount or not user_id:
                    return Response({'message': 'Body parameters not valid. Check and try again.'},
                                    status=status.HTTP_400_BAD_REQUEST)

                prices_dict = {
                    1000: 4.0,
                    2000: 7.9,
                    3000: 11.1,
                    4000: 14.6,
                    5000: 18.1,
                    6000: 21.1,
                    7000: 25.1,
                    8000: 27.1,
                    10000: 32.1,
                    15000: 48.1,
                    20000: 64.1,
                    25000: 78.1,
                    30000: 94.1,
                    40000: 128.1,
                    50000: 155.1,
                    100000: 290.1
                }

                amount_to_be_deducted = prices_dict[data_volume]
                if amount != amount_to_be_deducted:
                    return Response({'message': 'Invalid amount.'},
                                    status=status.HTTP_400_BAD_REQUEST)
                print(str(amount_to_be_deducted) + "================")
                # channel = phone_number
                date = datetime.datetime.now().strftime("%a, %b %d, %Y")
                time = datetime.datetime.now().strftime("%I:%M:%S %p")
                date_and_time = datetime.datetime.now().isoformat()
                if "wallet" == "wallet":
                    print("used this")
                    try:
                        enough_balance = check_user_balance_against_price(user_id, amount_to_be_deducted)
                    except:
                        return Response(
                            {'code': '0001', 'message': f'User ID does not exist: User ID provided: {user_id}.'},
                            status=status.HTTP_400_BAD_REQUEST)
                else:
                    enough_balance = True
                    print("not wallet")
                print(enough_balance)
                if enough_balance:
                    user_details = get_user_details(user_id)
                    first_name = user_details['first name']
                    last_name = user_details['last name']
                    email = user_details['email']
                    phone = user_details['phone']
                    bal = user_details['wallet']
                    # hist = history_web.collection(email).document(date_and_time)
                    # doc = hist.get()
                    # if doc.exists:
                    #     print(doc)
                    #     return redirect(f"https://{callback_url}")
                    # else:
                    #     print("no record found")
                    if "wallet" == "wallet":
                        print("updated")
                        user = get_user_details(user_id)
                        if user is None:
                            return None
                        previous_user_wallet = user['wallet']
                        print(f"previous wallet: {previous_user_wallet}")
                        new_balance = float(previous_user_wallet) - float(amount_to_be_deducted)
                        print(f"new_balance:{new_balance}")
                        doc_ref = user_collection.document(user_id)
                        doc_ref.update({'wallet': new_balance})
                        user = get_user_details(user_id)
                        new_user_wallet = user['wallet']
                        print(f"new_user_wallet: {new_user_wallet}")
                        if new_user_wallet == previous_user_wallet:
                            user = get_user_details(user_id)
                            if user is None:
                                return None
                            previous_user_wallet = user['wallet']
                            print(f"previous wallet: {previous_user_wallet}")
                            new_balance = float(previous_user_wallet) - float(amount_to_be_deducted)
                            print(f"new_balance:{new_balance}")
                            doc_ref = user_collection.document(user_id)
                            doc_ref.update({'wallet': new_balance})
                            user = get_user_details(user_id)
                            new_user_wallet = user['wallet']
                            print(f"new_user_wallet: {new_user_wallet}")
                        else:
                            print("it's fine")

                    data = {
                        'batch_id': "unknown",
                        'buyer': phone,
                        'color_code': "Green",
                        'amount': amount_to_be_deducted,
                        'data_break_down': data_volume,
                        'data_volume': data_volume,
                        'date': str(date),
                        'date_and_time': str(date_and_time),
                        'done': "unknown",
                        'email': email,
                        'image': user_id,
                        'ishareBalance': '',
                        'name': f"{first_name} {last_name}",
                        'number': receiver,
                        'paid_at': str(date_and_time),
                        'reference': reference,
                        'responseCode': 200,
                        'status': "Undelivered",
                        'bal': bal,
                        'time': str(time),
                        'tranxId': str(tranx_id_generator()),
                        'type': "MTN Master Bundle",
                        'uid': user_id
                    }

                    history_collection.document(date_and_time).set(data)
                    history_web.collection(email).document(date_and_time).set(data)
                    user = history_collection.document(date_and_time)
                    doc = user.get()
                    print(doc.to_dict())
                    tranx_id = doc.to_dict()['tranxId']
                    second_data = {
                        'amount': amount_to_be_deducted,
                        'batch_id': "unknown",
                        'channel': "wallet",
                        'color_code': "Green",
                        'created_at': date_and_time,
                        'data_volume': data_volume,
                        'date': str(date),
                        'email': email,
                        'date_and_time': date_and_time,
                        'image': user_id,
                        'ip_address': "",
                        'ishareBalance': 0,
                        'name': f"{first_name} {last_name}",
                        'number': receiver,
                        'buyer': phone,
                        'paid_at': date_and_time,
                        'payment_status': "success",
                        'reference': reference,
                        'status': "Undelivered",
                        'bal': bal,
                        'time': str(time),
                        'tranxId': tranx_id,
                        'type': "MTN Master Bundle"
                    }
                    mtn_other.document(date_and_time).set(second_data)
                    print("pu")

                    # tot = user_collection.document(user_id)
                    # print(tot.get().to_dict())
                    # try:
                    #     print(tot.get().to_dict()['mtn_total_sales'])
                    #     previous_sale = tot.get().to_dict()['mtn_total_sales']
                    #     print(f"Previous Sale: {previous_sale}")
                    #     new_sale = float(previous_sale) + float(amount_to_be_deducted)
                    #     print(new_sale)
                    #     user_collection.document(user_id).update({'mtn_total_sales': new_sale})
                    # except:
                    #     user_collection.document(user_id).update({'mtn_total_sales': amount_to_be_deducted})

                    #                     tat = cashback_collection.document(user_id)
                    #                     print(tat.get().to_dict())
                    #
                    #                     try:
                    #                         previous_cashback = tat.get().to_dict()['cashback_wallet']
                    #                         print(previous_cashback)
                    #                         cashback_balance = (0.5 / 100) * float(amount_to_be_deducted)
                    #                         print(cashback_balance)
                    #                         new_cashback = float(previous_cashback) + float(cashback_balance)
                    #                         print(new_cashback)
                    #                         cashback_collection.document(user_id).update(
                    #                             {'cashback_wallet': new_cashback, 'phone_number': user_details['phone']})
                    #
                    #                     except TypeError as e:
                    #                         print(e)
                    #                         cashback_balance = (0.5 / 100) * float(amount_to_be_deducted)
                    #                         print(cashback_balance)
                    #                         cashback_collection.document(user_id).set(
                    #                             {'cashback_wallet': cashback_balance, 'phone_number': user_details['phone']})
                    #
                    #                         print(cashback_collection.document(user_id).get().to_dict())
                    #                         print("did")

                    mail_doc_ref = mail_collection.document()
                    file_path = 'business_api/mtn_maill.txt'  # Replace with your file path

                    name = first_name
                    volume = data_volume
                    date = date_and_time
                    reference_t = reference
                    receiver_t = receiver

                    with open(file_path, 'r') as file:
                        html_content = file.read()

                    placeholders = {
                        '{name}': name,
                        '{volume}': volume,
                        '{date}': date,
                        '{reference}': reference_t,
                        '{receiver}': receiver_t
                    }

                    for placeholder, value in placeholders.items():
                        html_content = html_content.replace(placeholder, str(value))

                    mail_doc_ref.set({
                        'to': email,
                        'message': {
                            'subject': 'MTN Data',
                            'html': html_content,
                            'messageId': 'CloudHub GH'
                        }
                    })
                    print("got to redirect")
                    return Response(data={"status": "200", "message": "Transaction received successfully"},
                                    status=status.HTTP_200_OK)
                else:
                    return Response({"status": '400', 'message': 'Not enough balance to perform transaction'},
                                    status=status.HTTP_400_BAD_REQUEST)
            except Token.DoesNotExist:
                return Response({'error': 'Token does not exist.'}, status=status.HTTP_401_UNAUTHORIZED)
        else:
            return Response({'error': 'Invalid Header Provided.'}, status=status.HTTP_401_UNAUTHORIZED)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@authentication_classes([BearerTokenAuthentication])
def initiate_ishare_transaction(request):
    authorization_header = request.headers.get('Authorization')
    if authorization_header:
        auth_type, token = authorization_header.split(' ')
        if auth_type == 'Bearer':
            try:

                receiver = request.data.get('receiver')
                print(receiver)
                data_volume = request.data.get('data_volume')
                print(data_volume)
                reference = request.data.get('reference')
                amount = request.data.get('amount')
                # phone_number = request.data.get('phone_number')
                # channel = request.data.get('channel')
                # txn_type = request.data.get('txn_type')
                # txn_status = request.data.get('txn_status')
                # paid_at = request.data.get('paid_at')
                # ishare_balance = request.data.get('ishare_balance')
                # color_code = request.data.get('color_code')
                # data_break_down = request.data.get('data_break_down')
                # image = request.data.get('image')

                if not receiver or not data_volume or not reference or not amount:
                    return Response({'message': 'Body parameters not valid. Check and try again.'},
                                    status=status.HTTP_400_BAD_REQUEST)

                token_obj = Token.objects.get(key=token)
                user = token_obj.user
                user_id = user.user_id

                date = datetime.datetime.now().strftime("%a, %b %d, %Y")
                time = datetime.datetime.now().strftime("%I:%M:%S %p")
                date_and_time = datetime.datetime.now().isoformat()

                if "wallet" == "wallet":
                    try:
                        enough_balance = check_user_at_balance_against_price(user_id, data_volume)
                    except:
                        return Response(
                            {'code': '0001', 'message': f'User ID does not exist: User ID provided: {user_id}.'},
                            status=status.HTTP_400_BAD_REQUEST)
                else:
                    enough_balance = True
                    print("not wallet")
                print(enough_balance)
                if enough_balance:
                    user_details = get_user_details(user_id)
                    email = user_details['email']
                    print(enough_balance)
                    # hist = history_web.collection(email).document(date_and_time)
                    # doc = hist.get()
                    # if doc.exists:
                    #     return redirect(f"https://{callback_url}")
                    # else:
                    #     print("no record found")
                    if "wallet" == "wallet":
                        user = get_user_details(user_id)
                        if user is None:
                            return None
                        previous_user_wallet = user['at_balance']
                        print(f"previous wallet: {previous_user_wallet}")
                        new_balance = float(previous_user_wallet) - float(data_volume)
                        print(f"new_balance:{new_balance}")
                        doc_ref = user_collection.document(user_id)
                        doc_ref.update({'at_balance': new_balance})
                        user = get_user_details(user_id)
                        new_user_wallet = user['at_balance']
                        print(f"new_user_wallet: {new_user_wallet}")
                        if new_user_wallet == previous_user_wallet:
                            user = get_user_details(user_id)
                            if user is None:
                                return None
                            previous_user_wallet = user['at_balance']
                            print(f"previous wallet: {previous_user_wallet}")
                            new_balance = float(previous_user_wallet) - float(data_volume)
                            print(f"new_balance:{new_balance}")
                            doc_ref = user_collection.document(user_id)
                            doc_ref.update({'at_balance': new_balance})
                            user = get_user_details(user_id)
                            new_user_wallet = user['at_balance']
                            print(f"new_user_wallet: {new_user_wallet}")
                        else:
                            print("it's fine")
                    ishare_response = send_and_save_to_history(user_id, float(data_volume), reference,
                                                               float(amount), receiver,
                                                               date, time,
                                                               date_and_time)
                    print(ishare_response.status_code)
                    if ishare_response.status_code == 401:
                        return Response(
                            data={'status_code': ishare_response.status_code, "message": "Authorization Failed"},
                            status=status.HTTP_400_BAD_REQUEST)
                    data = ishare_response.json()
                    try:
                        print("entered the try")
                        batch_id = data["batchId"]
                        print("batch id")
                    except KeyError:
                        print("key error")
                        return Response(
                            data={'status_code': ishare_response.status_code, "message": "Transaction Failed"},
                            status=status.HTTP_400_BAD_REQUEST)
                    print(data["batchId"])
                    status_code = ishare_response.status_code
                    if batch_id is None:
                        print("batch id was none")
                        return Response(data={'status_code': status_code, "message": "Transaction Failed"},
                                        status=status.HTTP_400_BAD_REQUEST)

                    sms = f"Your account has been credited with {data_volume}MB."
                    r_sms_url = f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UmpEc1JzeFV4cERKTWxUWktqZEs&to={receiver}&from=CloudHub GH&sms={sms}"
                    response = requests.request("GET", url=r_sms_url)
                    print(response.text)
                    doc_ref = history_collection.document(date_and_time)
                    doc_ref.update({'done': 'Successful'})
                    mail_doc_ref = mail_collection.document(f"{batch_id}-Mail")
                    file_path = 'business_api/mail.txt'  # Replace with your file path

                    name = user["first name"]
                    volume = data_volume
                    date = date_and_time
                    reference_t = reference
                    receiver_t = receiver

                    with open(file_path, 'r') as file:
                        html_content = file.read()

                    placeholders = {
                        '{name}': name,
                        '{volume}': volume,
                        '{date}': date,
                        '{reference}': reference_t,
                        '{receiver}': receiver_t
                    }

                    for placeholder, value in placeholders.items():
                        html_content = html_content.replace(placeholder, str(value))

                    mail_doc_ref.set({
                        'to': email,
                        'message': {
                            'subject': 'AT Flexi Bundle',
                            'html': html_content,
                            'messageId': 'CloudHub GH'
                        }
                    })

                    # tot = user_collection.document(user_id)
                    # print(tot.get().to_dict())
                    # try:
                    #     print(tot.get().to_dict()['at_total_sales'])
                    #     previous_sale = tot.get().to_dict()['at_total_sales']
                    #     print(f"Previous Sale: {previous_sale}")
                    #     new_sale = float(previous_sale) + float(amount)
                    #     print(new_sale)
                    #     user_collection.document(user_id).update({'at_total_sales': new_sale})
                    # except:
                    #     user_collection.document(user_id).update({'at_total_sales': amount})

                    # tat = cashback_collection.document(user_id)
                    # print(tat.get().to_dict())
                    #
                    # try:
                    #     previous_cashback = tat.get().to_dict()['cashback_wallet']
                    #     print(previous_cashback)
                    #     cashback_balance = (0.5 / 100) * float(amount)
                    #     print(cashback_balance)
                    #     new_cashback = float(previous_cashback) + float(cashback_balance)
                    #     print(new_cashback)
                    #     cashback_collection.document(user_id).update(
                    #         {'cashback_wallet': new_cashback, 'phone_number': user_details['phone']})
                    #
                    # except TypeError as e:
                    #     print(e)
                    #     cashback_balance = (0.5 / 100) * float(amount)
                    #     print(cashback_balance)
                    #     cashback_collection.document(user_id).set(
                    #         {'cashback_wallet': cashback_balance, 'phone_number': user_details['phone']})
                    #
                    #     print(cashback_collection.document(user_id).get().to_dict())
                    #     print("did")

                    return Response(data={'status_code': status_code, 'batch_id': batch_id},
                                    status=status.HTTP_200_OK)
                else:
                    return Response(data={'status_code': 400, "message": "Not enough balance"},
                                    status=status.HTTP_400_BAD_REQUEST)
            except Token.DoesNotExist:
                return Response({'error': 'Token does not exist.'}, status=status.HTTP_401_UNAUTHORIZED)
        else:
            return Response({'error': 'Invalid Header Provided.'}, status=status.HTTP_401_UNAUTHORIZED)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@authentication_classes([BearerTokenAuthentication])
def admin_initiate_ishare_transaction(request):
    # allowed_hosts = ['cloudhubgh.com', 'reseller.cloudhubgh.com', "api.cloudhubgh.com", "merchant.cloudhubgh.com"]
    # request_host = request.headers.get('Host')
    #
    # print(f"hosssssssssstttttttttttttttt issssssssssssss {request_host}")
    #
    # if request_host not in allowed_hosts:
    #     response1 = requests.get(
    #         f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UmpEc1JzeFV4cERKTWxUWktqZEs&to=0592117523&from=Bundle&sms=Invalid header host={request_host}")
    #     print(response1.text)
    #     return JsonResponse({'error': 'Host not allowed'}, status=403)
    allowed_origins = ['https://reseller.cloudhubgh.com']

    if 'HTTP_ORIGIN' not in request.META:
        return JsonResponse({'error': 'Origin header missing'}, status=400)

    request_origin = request.META['HTTP_ORIGIN']
    print(f"origiiiiiiiiiiiiiiinnnnnnnnnnnnnnnnn issssssssssssssssssssssssssssssssssssssssss {request_origin}")

    if request_origin not in allowed_origins:
        return JsonResponse({'error': 'Origin not allowed'}, status=403)
    authorization_header = request.headers.get('Authorization')
    if authorization_header:
        auth_type, token = authorization_header.split(' ')
        if auth_type == 'Bearer':
            try:
                receiver = request.data.get('receiver')
                print(receiver)
                data_volume = request.data.get('data_volume')
                print(data_volume)
                reference = request.data.get('reference')
                amount = request.data.get('amount')
                user_id = request.data.get('user_id')
                # phone_number = request.data.get('phone_number')
                # channel = request.data.get('channel')
                # txn_type = request.data.get('txn_type')
                # txn_status = request.data.get('txn_status')
                # paid_at = request.data.get('paid_at')
                # ishare_balance = request.data.get('ishare_balance')
                # color_code = request.data.get('color_code')
                # data_break_down = request.data.get('data_break_down')
                # image = request.data.get('image')

                if not receiver or not data_volume or not reference or not amount:
                    print("body parameters")
                    return Response({'message': 'Body parameters not valid. Check and try again.'},
                                    status=status.HTTP_400_BAD_REQUEST)

                token_obj = Token.objects.get(key=token)
                token_key = token_obj.key

                if token_key != config("TOKEN_KEY"):
                    print("token did not match")
                    return Response({'message': 'Authorisation Failed.'},
                                    status=status.HTTP_401_UNAUTHORIZED)

                date = datetime.datetime.now().strftime("%a, %b %d, %Y")
                time = datetime.datetime.now().strftime("%I:%M:%S %p")
                date_and_time = datetime.datetime.now().isoformat()

                if "wallet" == "wallet":
                    try:
                        enough_balance = check_user_balance_against_price(user_id, amount)
                    except:
                        return Response(
                            {'code': '0001', 'message': f'User ID does not exist: User ID provided: {user_id}.'},
                            status=status.HTTP_400_BAD_REQUEST)
                else:
                    enough_balance = True
                    print("not wallet")
                print(enough_balance)
                if enough_balance:
                    user_details = get_user_details(user_id)
                    email = user_details['email']
                    print(enough_balance)
                    # hist = history_web.collection(email).document(date_and_time)
                    # doc = hist.get()
                    # if doc.exists:
                    #     return redirect(f"https://{callback_url}")
                    # else:
                    #     print("no record found")
                    if "wallet" == "wallet":
                        user = get_user_details(user_id)
                        if user is None:
                            return None
                        previous_user_wallet = user['wallet']
                        print(f"previous wallet: {previous_user_wallet}")
                        new_balance = float(previous_user_wallet) - float(amount)
                        print(f"new_balance:{new_balance}")
                        doc_ref = user_collection.document(user_id)
                        doc_ref.update({'wallet': new_balance})
                        user = get_user_details(user_id)
                        new_user_wallet = user['wallet']
                        print(f"new_user_wallet: {new_user_wallet}")
                        if new_user_wallet == previous_user_wallet:
                            user = get_user_details(user_id)
                            if user is None:
                                return None
                            previous_user_wallet = user['wallet']
                            print(f"previous wallet: {previous_user_wallet}")
                            new_balance = float(previous_user_wallet) - float(amount)
                            print(f"new_balance:{new_balance}")
                            doc_ref = user_collection.document(user_id)
                            doc_ref.update({'wallet': new_balance})
                            user = get_user_details(user_id)
                            new_user_wallet = user['wallet']
                            print(f"new_user_wallet: {new_user_wallet}")
                        else:
                            print("it's fine")
                    ishare_response = send_and_save_to_history(user_id, float(data_volume), reference,
                                                               float(amount), receiver,
                                                               date, time,
                                                               date_and_time)
                    print(ishare_response.status_code)
                    if ishare_response.status_code == 401:
                        return Response(
                            data={'status_code': ishare_response.status_code, "message": "Authorization Failed"},
                            status=status.HTTP_400_BAD_REQUEST)
                    data = ishare_response.json()
                    try:
                        print("entered the try")
                        batch_id = data["batchId"]
                        print("batch id")
                    except KeyError:
                        print("key error")
                        return Response(
                            data={'status_code': ishare_response.status_code, "message": "Transaction Failed"},
                            status=status.HTTP_400_BAD_REQUEST)
                    print(data["batchId"])
                    status_code = ishare_response.status_code
                    print(f"status codeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeee: {status_code}")
                    if batch_id is None:
                        print("batch id was none")
                        return Response(data={'status_code': status_code, "message": "Transaction Failed"},
                                        status=status.HTTP_400_BAD_REQUEST)

                    sms = f"Your account has been credited with {data_volume}MB."
                    r_sms_url = f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UmpEc1JzeFV4cERKTWxUWktqZEs&to={receiver}&from=Bundle&sms={sms}"
                    response = requests.request("GET", url=r_sms_url)
                    print(response.text)
                    doc_ref = history_collection.document(date_and_time)
                    doc_ref.update({'done': 'Successful'})
                    mail_doc_ref = mail_collection.document(f"{batch_id}-Mail")
                    file_path = 'business_api/mail.txt'  # Replace with your file path

                    name = user_details["first name"]
                    volume = data_volume
                    date = date_and_time
                    reference_t = reference
                    receiver_t = receiver

                    with open(file_path, 'r') as file:
                        html_content = file.read()

                    placeholders = {
                        '{name}': name,
                        '{volume}': volume,
                        '{date}': date,
                        '{reference}': reference_t,
                        '{receiver}': receiver_t
                    }

                    for placeholder, value in placeholders.items():
                        html_content = html_content.replace(placeholder, str(value))

                    mail_doc_ref.set({
                        'to': email,
                        'message': {
                            'subject': 'AT Flexi Bundle',
                            'html': html_content,
                            'messageId': 'CloudHub GH'
                        }
                    })

                    # tot = user_collection.document(user_id)
                    # print(tot.get().to_dict())
                    # try:
                    #     print(tot.get().to_dict()['at_total_sales'])
                    #     previous_sale = tot.get().to_dict()['at_total_sales']
                    #     print(f"Previous Sale: {previous_sale}")
                    #     new_sale = float(previous_sale) + float(amount)
                    #     print(new_sale)
                    #     user_collection.document(user_id).update({'at_total_sales': new_sale})
                    # except:
                    #     user_collection.document(user_id).update({'at_total_sales': amount})

                    #                     tat = cashback_collection.document(user_id)
                    #                     print(tat.get().to_dict())
                    #
                    #                     try:
                    #                         previous_cashback = tat.get().to_dict()['cashback_wallet']
                    #                         print(previous_cashback)
                    #                         cashback_balance = (0.5 / 100) * float(amount)
                    #                         print(cashback_balance)
                    #                         new_cashback = float(previous_cashback) + float(cashback_balance)
                    #                         print(new_cashback)
                    #                         cashback_collection.document(user_id).update(
                    #                             {'cashback_wallet': new_cashback, 'phone_number': user_details['phone']})
                    #
                    #                     except TypeError as e:
                    #                         print(e)
                    #                         cashback_balance = (0.5 / 100) * float(amount)
                    #                         print(cashback_balance)
                    #                         cashback_collection.document(user_id).set(
                    #                             {'cashback_wallet': cashback_balance, 'phone_number': user_details['phone']})
                    #
                    #                         print(cashback_collection.document(user_id).get().to_dict())
                    #                         print("did")

                    return Response(data={'status_code': status_code, 'batch_id': batch_id},
                                    status=status.HTTP_200_OK)
                else:
                    return Response(data={'status_code': 400, "message": "Not enough balance"},
                                    status=status.HTTP_400_BAD_REQUEST)
            except Token.DoesNotExist:
                return Response({'error': 'Token does not exist.'}, status=status.HTTP_401_UNAUTHORIZED)
        else:
            print("hereeeeeeeeee")
            return Response({'error': 'Invalid Header Provided.'}, status=status.HTTP_401_UNAUTHORIZED)


@api_view(['GET'])
@authentication_classes([])
@permission_classes([AllowAny])
def get_user_token(request):
    user_id = request.data.get('user_id')
    if user_id:
        try:
            user = models.CustomUser.objects.get(user_id=user_id)
            token = Token.objects.get(user=user)
            return Response({'user_id': user_id, 'token': token.key}, status=status.HTTP_200_OK)
        except models.CustomUser.DoesNotExist:
            return Response({'message': 'User does not exist.'}, status=status.HTTP_404_NOT_FOUND)
        except Token.DoesNotExist:
            return Response({'message': 'Token does not exist for the user.'}, status=status.HTTP_404_NOT_FOUND)
    else:
        return Response({'message': 'User ID parameter is missing.'}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@authentication_classes([BearerTokenAuthentication])
def initiate_big_time(request):
    authorization_header = request.headers.get('Authorization')
    if authorization_header:
        auth_type, token = authorization_header.split(' ')
        if auth_type == 'Bearer':
            try:
                token_obj = Token.objects.get(key=token)
                user = token_obj.user
                user_id = user.user_id
                print(user_id)

                print("hiiiii")

                prices_dict = {
                    30000: 80,
                    40000: 100,
                    50000: 120,
                    80000: 200,
                    100000: 230,
                    200000: 450,
                }

                receiver = request.data.get('receiver')
                print(receiver)
                data_volume = request.data.get('data_volume')
                reference = request.data.get('reference')
                print(data_volume, reference)
                try:
                    amount = prices_dict[data_volume]
                    print(amount)
                except KeyError:
                    print("key error")
                    return Response({'message': 'Check data volume parameter and try again.'},
                                    status=status.HTTP_400_BAD_REQUEST)
                print(amount)
                # phone_number = request.data.get('phone_number')

                print("yo")

                if "wallet" == "wallet":
                    print("used this")
                    try:
                        enough_balance = check_user_balance_against_price(user_id, amount)
                    except:
                        return Response(
                            {'code': '0001', 'message': f'User ID does not exist: User ID provided: {user_id}.'},
                            status=status.HTTP_400_BAD_REQUEST)
                else:
                    enough_balance = True
                    print("not wallet")
                print(enough_balance)
                if enough_balance:

                    if not receiver or not data_volume or not reference or not amount:
                        return Response({'message': 'Body parameters not valid. Check and try again.'},
                                        status=status.HTTP_400_BAD_REQUEST)

                    print("got here")
                    user_details = get_user_details(user_id)
                    print(user_details['first name'])

                    date = datetime.datetime.now().strftime("%a, %b %d, %Y")
                    time = datetime.datetime.now().strftime("%I:%M:%S %p")
                    date_and_time = datetime.datetime.now().isoformat()

                    if user_details is not None:
                        print("yes")
                        first_name = user_details['first name']
                        print(first_name)
                        last_name = user_details['last name']
                        print(last_name)
                        email = user_details['email']
                        phone = user_details['phone']
                    else:
                        first_name = ""
                        last_name = ""
                        email = ""
                        phone = ""
                    details = {
                        'first_name': first_name,
                        'last_name': last_name,
                        'email': email,
                        'user_id': user_id
                    }
                    big_time_response = big_time_transaction(receiver=receiver, date_and_time=date_and_time, date=date,
                                                             time=time, amount=amount, data_volume=data_volume,
                                                             channel="MoMo", phone=phone, ref=reference,
                                                             details=details, txn_status="Undelivered", user_id=user_id)
                    if big_time_response.status_code == 200 or big_time_response.data["code"] == "0000":
                        if "wallet" == "wallet":
                            print("updated")
                            user = get_user_details(user_id)
                            if user is None:
                                return None
                            previous_user_wallet = user['wallet']
                            print(f"previous wallet: {previous_user_wallet}")
                            new_balance = float(previous_user_wallet) - float(amount)
                            print(f"new_balance:{new_balance}")
                            doc_ref = user_collection.document(user_id)
                            doc_ref.update({'wallet': new_balance})
                            user = get_user_details(user_id)
                            new_user_wallet = user['wallet']
                            print(f"new_user_wallet: {new_user_wallet}")
                            if new_user_wallet == previous_user_wallet:
                                user = get_user_details(user_id)
                                if user is None:
                                    return None
                                previous_user_wallet = user['wallet']
                                print(f"previous wallet: {previous_user_wallet}")
                                new_balance = float(previous_user_wallet) - float(amount)
                                print(f"new_balance:{new_balance}")
                                doc_ref = user_collection.document(user_id)
                                doc_ref.update({'wallet': new_balance})
                                user = get_user_details(user_id)
                                new_user_wallet = user['wallet']
                                print(f"new_user_wallet: {new_user_wallet}")
                            else:
                                print("it's fine")
                        return Response(data={"status": "200", "message": "Transaction received successfully"},
                                        status=status.HTTP_200_OK)
                    else:
                        return Response({"status": 400, "message": "Insufficient balance"},
                                        status=status.HTTP_400_BAD_REQUEST)
                else:
                    return Response({"status": '400', 'message': 'Something went wrong'},
                                    status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                print(e)
                return Response({"status": '400', 'message': f'Something went wrong: {e}'},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'error': 'Invalid Header Provided.'}, status=status.HTTP_401_UNAUTHORIZED)
    else:
        return Response({'error': 'Invalid Header Provided.'}, status=status.HTTP_401_UNAUTHORIZED)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@authentication_classes([BearerTokenAuthentication])
def admin_initiate_big_time(request):
    # allowed_hosts = ['cloudhubgh.com', 'reseller.cloudhubgh.com', "api.cloudhubgh.com", "merchant.cloudhubgh.com"]
    # request_host = request.headers.get('Host')
    #
    # print(f"hosssssssssstttttttttttttttt issssssssssssss {request_host}")
    #
    # if request_host not in allowed_hosts:
    #     response1 = requests.get(
    #         f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UmpEc1JzeFV4cERKTWxUWktqZEs&to=0592117523&from=Bundle&sms=Invalid header host={request_host}")
    #     print(response1.text)
    #     return JsonResponse({'error': 'Host not allowed'}, status=403)
    allowed_origins = ['https://reseller.cloudhubgh.com']

    if 'HTTP_ORIGIN' not in request.META:
        return JsonResponse({'error': 'Origin header missing'}, status=400)

    request_origin = request.META['HTTP_ORIGIN']
    print(f"origiiiiiiiiiiiiiiinnnnnnnnnnnnnnnnn issssssssssssssssssssssssssssssssssssssssss {request_origin}")

    if request_origin not in allowed_origins:
        return JsonResponse({'error': 'Origin not allowed'}, status=403)
    authorization_header = request.headers.get('Authorization')
    if authorization_header:
        auth_type, token = authorization_header.split(' ')
        if auth_type == 'Bearer':
            try:
                token_obj = Token.objects.get(key=token)
                user = token_obj.user

                print("hiiiii")

                prices_dict = {
                    30000: 80,
                    40000: 100,
                    50000: 120,
                    80000: 200,
                    100000: 230,
                    200000: 450,
                }

                receiver = request.data.get('receiver')
                print(receiver)
                data_volume = request.data.get('data_volume')
                reference = request.data.get('reference')
                user_id = request.data.get('user_id')
                print(data_volume, reference)
                try:
                    amount = prices_dict[data_volume]
                    if amount != amount:
                        return Response({'message': 'Invalid amount.'},
                                        status=status.HTTP_400_BAD_REQUEST)
                    print(amount)
                except KeyError:
                    print("key error")
                    return Response({'message': 'Check data volume parameter and try again.'},
                                    status=status.HTTP_400_BAD_REQUEST)
                print(amount)
                # phone_number = request.data.get('phone_number')

                print("yo")

                if "wallet" == "wallet":
                    print("used this")
                    try:
                        enough_balance = check_user_balance_against_price(user_id, amount)
                    except:
                        return Response(
                            {'code': '0001', 'message': f'User ID does not exist: User ID provided: {user_id}.'},
                            status=status.HTTP_400_BAD_REQUEST)
                else:
                    enough_balance = True
                    print("not wallet")
                print(enough_balance)
                if enough_balance:

                    if not receiver or not data_volume or not reference or not amount:
                        return Response({'message': 'Body parameters not valid. Check and try again.'},
                                        status=status.HTTP_400_BAD_REQUEST)

                    token_obj = Token.objects.get(key=token)
                    token_key = token_obj.key

                    if token_key != config("TOKEN_KEY"):
                        return Response({'message': 'Authorisation Failed.'},
                                        status=status.HTTP_400_BAD_REQUEST)

                    print("got here")
                    user_details = get_user_details(user_id)
                    print(user_details['first name'])

                    date = datetime.datetime.now().strftime("%a, %b %d, %Y")
                    time = datetime.datetime.now().strftime("%I:%M:%S %p")
                    date_and_time = datetime.datetime.now().isoformat()

                    if user_details is not None:
                        print("yes")
                        first_name = user_details['first name']
                        print(first_name)
                        last_name = user_details['last name']
                        print(last_name)
                        email = user_details['email']
                        phone = user_details['phone']
                    else:
                        first_name = ""
                        last_name = ""
                        email = ""
                        phone = ""
                    details = {
                        'first_name': first_name,
                        'last_name': last_name,
                        'email': email,
                        'user_id': user_id
                    }
                    big_time_response = big_time_transaction(receiver=receiver, date_and_time=date_and_time, date=date,
                                                             time=time, amount=amount, data_volume=data_volume,
                                                             channel="MoMo", phone=phone, ref=reference,
                                                             details=details, txn_status="Undelivered", user_id=user_id)
                    if big_time_response.status_code == 200 or big_time_response.data["code"] == "0000":
                        if "wallet" == "wallet":
                            print("updated")
                            user = get_user_details(user_id)
                            if user is None:
                                return None
                            previous_user_wallet = user['wallet']
                            print(f"previous wallet: {previous_user_wallet}")
                            new_balance = float(previous_user_wallet) - float(amount)
                            print(f"new_balance:{new_balance}")
                            doc_ref = user_collection.document(user_id)
                            doc_ref.update({'wallet': new_balance})
                            user = get_user_details(user_id)
                            new_user_wallet = user['wallet']
                            print(f"new_user_wallet: {new_user_wallet}")
                            if new_user_wallet == previous_user_wallet:
                                user = get_user_details(user_id)
                                if user is None:
                                    return None
                                previous_user_wallet = user['wallet']
                                print(f"previous wallet: {previous_user_wallet}")
                                new_balance = float(previous_user_wallet) - float(amount)
                                print(f"new_balance:{new_balance}")
                                doc_ref = user_collection.document(user_id)
                                doc_ref.update({'wallet': new_balance})
                                user = get_user_details(user_id)
                                new_user_wallet = user['wallet']
                                print(f"new_user_wallet: {new_user_wallet}")
                            else:
                                print("it's fine")
                        return Response(data={"status": "200", "message": "Transaction received successfully"},
                                        status=status.HTTP_200_OK)
                    else:
                        return Response({"status": 400, "message": "Something went wrong"},
                                        status=status.HTTP_400_BAD_REQUEST)
                else:
                    return Response({"status": 400, 'message': 'Insufficient Balance'},
                                    status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                print(e)
                return Response({"status": '400', 'message': f'Something went wrong: {e}'},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'error': 'Invalid Header Provided.'}, status=status.HTTP_401_UNAUTHORIZED)
    else:
        return Response({'error': 'Invalid Header Provided.'}, status=status.HTTP_401_UNAUTHORIZED)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@authentication_classes([BearerTokenAuthentication])
def wallet_topup(request):
    allowed_origins = ['https://reseller.cloudhubgh.com']

    if 'HTTP_ORIGIN' not in request.META:
        return JsonResponse({'error': 'Origin header missing'}, status=400)

    request_origin = request.META['HTTP_ORIGIN']
    print(f"origiiiiiiiiiiiiiiinnnnnnnnnnnnnnnnn issssssssssssssssssssssssssssssssssssssssss {request_origin}")

    if request_origin not in allowed_origins:
        return JsonResponse({'error': 'Origin not allowed'}, status=403)
    authorization_header = request.headers.get('Authorization')
    if authorization_header:
        auth_type, token = authorization_header.split(' ')
        if auth_type == 'Bearer':
            try:
                print(
                    "****************************************************************************************************")
                token_obj = Token.objects.get(key=token)
                user = token_obj.user
                user_id = user.user_id

                amount = request.data.get('topup_amount')
                reference = request.data.get('reference')
                receiver_id = request.data.get('receiver_id')
                user_id = request.data.get("user_id")

                allowed_ids = [
                    "eq5bk7lBvKUiXvJhY67B2Dd96RZ2", "xxWAltejVRU9eBHS3wtvL3HTbmE2", "9VA0qyq6lXYPZ6Ut867TVcBvF2t1",
                    "eq5bk7lBvKUiXvJhY67B2Dd96RZ2", "YvOWjIt9arboXcrOtxXEjbiVpZy1", "ajoExmDwsmQfkDlf1junWVU1SEt2",
                    "7khtAAe7awOlRqOg3YYb1f7VTeE3", "Q0hDkakaiTMDuvbYsD65mYE9W773", "13xmcsrIiDdVu1EDEp23NBT6QYj1",
                    "c0jyJW4MeqcEJ9LuJqhc7Rj4c9c2", "E4hsRsONxxNv27cQxNrIVC6xtL63", "7NvMDSUG4XcIzSNYs6LktMdo3X62",
                    "viRbi1VheufDqYJe2UGpy5WB3fA3", "jkzLTRl5cSWwIRdWJYCJFOK8iln1", "PqevIUYHlBhDoz6lpahssMjtqYH2",
                    "lW57LDVdj9O6IV3gwfNoI3gCELN2"
                ]

                if user_id not in allowed_ids:
                    return Response({'message': 'Not permitted.'},
                                    status=status.HTTP_400_BAD_REQUEST)

                print(
                    "==================================================================================================")
                print(
                    "==================================================================================================")
                print(
                    "==================================================================================================")
                print(amount, reference, receiver_id, user_id)

                if not amount or not reference:
                    return Response({'message': 'Body parameters not valid. Check and try again.'},
                                    status=status.HTTP_400_BAD_REQUEST)

                date = datetime.datetime.now().strftime("%a, %b %d, %Y")
                time = datetime.datetime.now().strftime("%I:%M:%S %p")
                date_and_time = datetime.datetime.now().isoformat()

                token_key = token_obj.key

                if token_key != config("TOKEN_KEY"):
                    return Response({'message': 'Authorisation Failed.'},
                                    status=status.HTTP_400_BAD_REQUEST)

                user_details = get_user_details(user_id)
                receiver_details = get_user_details(receiver_id)

                if user_details is not None:
                    print(user_details)
                    first_name = user_details['first name']
                    last_name = user_details['last name']
                    email = user_details['email']
                    phone = user_details['phone']
                    try:
                        previous_user_wallet = user_details['wallet']
                    except KeyError:
                        previous_user_wallet = 0

                    if float(previous_user_wallet) < float(amount):
                        return Response({'message': 'Insufficient Balance.'},
                                        status=status.HTTP_200_OK)

                    try:
                        previous_receiver_wallet = receiver_details['wallet']
                    except KeyError:
                        previous_receiver_wallet = 0
                else:
                    first_name = ""
                    last_name = ""
                    email = ""
                    phone = ""
                    previous_wallet = 0
                all_data = {
                    'batch_id': "unknown",
                    'buyer': user_details['phone'],
                    'color_code': "Green",
                    'amount': amount,
                    'data_break_down': amount,
                    'data_volume': amount,
                    'date': date,
                    'date_and_time': date_and_time,
                    'done': "Success",
                    'email': email,
                    'image': user_id,
                    'ishareBalance': 0,
                    'name': f"{first_name} {last_name}",
                    'number': receiver_details['phone'],
                    'paid_at': date_and_time,
                    'reference': reference,
                    'responseCode': 200,
                    'status': "Credited",
                    'time': time,
                    'tranxId': str(tranx_id_generator()),
                    'type': "WALLETTOPUP",
                    'uid': user_id
                }
                history_web.collection(email).document(date_and_time).set(all_data)
                print("f saved")
                history_collection.document(date_and_time).set(all_data)
                print(f"ya{history_collection.document(date_and_time).get().to_dict()}")
                print("f saved")
                print(f"yo{history_web.collection(email).document(date_and_time).get().to_dict()}")
                to_be_added = float(amount)
                to_be_deducted = float(amount)

                print("######################### got here ###########################################################")

                print(f"amount to be added: {to_be_added}")

                new_balance_for_receiver = previous_receiver_wallet + to_be_added
                print(f" new balance: {new_balance_for_receiver}")
                receiver_doc_ref = user_collection.document(receiver_id)
                receiver_doc_ref.update(
                    {'wallet': new_balance_for_receiver, 'wallet_last_update': date_and_time,
                     'recent_wallet_reference': reference})
                print(receiver_doc_ref.get().to_dict())
                print("before all data")

                print("######################### then here ###########################################################")

                new_balance_for_user = previous_user_wallet - to_be_deducted
                print(f" new balance: {new_balance_for_user}")
                user_doc_ref = user_collection.document(user_id)
                user_doc_ref.update(
                    {'wallet': new_balance_for_user, 'wallet_last_update': date_and_time,
                     'recent_wallet_reference': reference})
                print(receiver_doc_ref.get().to_dict())
                print("got to all data")
                all_data = {
                    'batch_id': "unknown",
                    'buyer': user_details['phone'],
                    'color_code': "Green",
                    'amount': amount,
                    'data_break_down': amount,
                    'data_volume': amount,
                    'date': date,
                    'date_and_time': date_and_time,
                    'done': "Success",
                    'email': email,
                    'image': user_id,
                    'ishareBalance': 0,
                    'name': f"{first_name} {last_name}",
                    'number': receiver_details['phone'],
                    'paid_at': date_and_time,
                    'reference': reference,
                    'responseCode': 200,
                    'status': "Credited",
                    'time': time,
                    'tranxId': str(tranx_id_generator()),
                    'type': "WALLETTOPUP",
                    'uid': user_id
                }
                print(
                    "savingggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggggg")
                history_web.collection(email).document(date_and_time).set(all_data)
                print("saved")
                history_collection.document(date_and_time).set(all_data)
                print(f"ya{history_collection.document(date_and_time).get().to_dict()}")
                print("saved")
                print(f"yo{history_web.collection(email).document(date_and_time).get().to_dict()}")
                print(
                    "savedddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddd")
                # name = f"{first_name} {last_name}"
                # amount = to_be_added
                # file_path = 'business_api/wallet_mail.txt'
                # mail_doc_ref = mail_collection.document()
                #
                # with open(file_path, 'r') as file:
                #     html_content = file.read()
                #
                # placeholders = {
                #     '{name}': name,
                #     '{amount}': amount
                # }
                #
                # for placeholder, value in placeholders.items():
                #     html_content = html_content.replace(placeholder, str(value))
                #
                # mail_doc_ref.set({
                #     'to': email,
                #     'message': {
                #         'subject': 'Wallet Topup',
                #         'html': html_content,
                #         'messageId': 'CloudHub GH'
                #     }
                # })

                sms_message = f"GHS {to_be_added} was deposited in your wallet. Available balance is now GHS {round(new_balance_for_receiver, 2)}"
                sms_url = f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UmpEc1JzeFV4cERKTWxUWktqZEs&to=0{receiver_details['phone']}&from=CloudHub GH&sms={sms_message}"
                response = requests.request("GET", url=sms_url)
                print(response.status_code)
                return Response(data={"status": "200", "message": "Wallet Topup Successful"},
                                status=status.HTTP_200_OK)
            except Token.DoesNotExist or models.CustomUser.DoesNotExist:
                return Response({'error': 'Token does not exist.'}, status=status.HTTP_401_UNAUTHORIZED)
        else:
            return Response({'error': 'Invalid Header Provided.'}, status=status.HTTP_401_UNAUTHORIZED)


def webhook_send_and_save_to_history(user_id, txn_type: str, paid_at: str, ishare_balance: float,
                                     color_code: str,
                                     data_volume: float, reference: str, amount: float,
                                     receiver: str,
                                     date: str, time: str, date_and_time: str, txn_status):
    user_details = get_user_details(user_id)
    first_name = user_details['first name']
    last_name = user_details['last name']
    email = user_details['email']
    phone = user_details['phone']

    doc_ref = history_web.collection(email).document(date_and_time)
    if doc_ref.get().exists:
        data = doc_ref.get()
        data_dict = data.to_dict()
        batch_id = data_dict["batch_id"]
        if batch_id != "unknown":
            ishare_verification_response = ishare_verification(batch_id)
            if ishare_verification_response is not False:
                code = \
                    ishare_verification_response["flexiIshareTranxStatus"]["flexiIshareTranxStatusResult"][
                        "apiResponse"][
                        "responseCode"]
                ishare_response = \
                    ishare_verification_response["flexiIshareTranxStatus"]["flexiIshareTranxStatusResult"][
                        "ishareApiResponseData"][
                        "apiResponseData"][
                        0][
                        "responseMsg"]
                print(code)
                print(ishare_response)
                if code == '200' or ishare_response == 'Crediting Successful.':
                    return HttpResponse(status=200)
                else:
                    pass
        else:
            pass
    else:
        pass
    print("moving on")
    data = {
        'batch_id': "unknown",
        'buyer': phone,
        'color_code': color_code,
        'amount': amount,
        'data_break_down': str(data_volume),
        'data_volume': data_volume,
        'date': date,
        'date_and_time': date_and_time,
        'done': "Pending",
        'email': email,
        'image': user_id,
        'ishareBalance': ishare_balance,
        'name': f"{first_name} {last_name}",
        'number': receiver,
        'paid_at': paid_at,
        'reference': reference,
        'responseCode': "0",
        'status': txn_status,
        'time': time,
        'tranxId': str(tranx_id_generator()),
        'type': txn_type,
        'uid': user_id
    }
    history_collection.document(date_and_time).set(data)
    history_web.collection(email).document(date_and_time).set(data)

    if history_collection.document(date_and_time).get().exists:
        print("first save")

    ishare_response = send_ishare_bundle(first_name=first_name, last_name=last_name, receiver=receiver,
                                         buyer=phone,
                                         bundle=data_volume,
                                         email=email)
    json_response = ishare_response.json()
    print(
        f"hello:{json_response}========================================================================================")
    status_code = ishare_response.status_code
    print(status_code)
    try:
        batch_id = json_response["batchId"]
    except KeyError:
        batch_id = "unknown"
    print(batch_id)

    doc_ref = history_collection.document(date_and_time)
    if doc_ref.get().exists:
        doc_ref.update({'batch_id': batch_id, 'responseCode': status_code})
        history_web.collection(email).document(date_and_time).update(
            {'batch_id': batch_id, 'responseCode': status_code})
    else:
        print("didn't find any entry to update")
    print("firebase saved")
    # return status_code, batch_id if batch_id else "No batchId", email, first_name
    return ishare_response


def mtn_flexi_transaction(receiver, date, time, date_and_time, phone, amount, data_volume, details: dict, ref,
                          channel, txn_status):
    data = {
        'batch_id': "unknown",
        'buyer': phone,
        'color_code': "Green",
        'amount': amount,
        'data_break_down': str(data_volume),
        'data_volume': data_volume,
        'date': date,
        'date_and_time': date_and_time,
        'done': "unknown",
        'email': details["email"],
        'image': details["user_id"],
        'ishareBalance': 0,
        'name': f"{details['first_name']} {details['last_name']}",
        'number': receiver,
        'paid_at': date_and_time,
        'reference': ref,
        'responseCode': 200,
        'status': txn_status,
        'time': time,
        'tranxId': str(tranx_id_generator()),
        'type': "MTN Master Bundle",
        'uid': details["user_id"]
    }

    history_collection.document(date_and_time).set(data)
    history_web.collection(details['email']).document(date_and_time).set(data)
    user = history_collection.document(date_and_time)
    # new_mtn_txn = models.MTNTransaction.objects.create(
    #     user_id=details["user_id"],
    #     amount=amount,
    #     bundle_volume=data_volume,
    #     number=receiver,
    #     firebase_date=date_and_time
    # )
    # new_mtn_txn.save()
    doc = user.get()
    print(doc.to_dict())
    tranx_id = doc.to_dict()['tranxId']
    second_data = {
        'amount': amount,
        'batch_id': "unknown",
        'channel': channel,
        'color_code': "Green",
        'created_at': date_and_time,
        'data_volume': data_volume,
        'date': date,
        'email': details["email"],
        'date_and_time': date_and_time,
        'image': details["user_id"],
        'ip_address': "",
        'ishareBalance': 0,
        'name': f"{details['first_name']} {details['last_name']}",
        'number': receiver,
        'buyer': phone,
        'paid_at': date_and_time,
        'payment_status': "success",
        'reference': ref,
        'status': txn_status,
        'time': time,
        'tranxId': tranx_id,
        'type': "MTN Master Bundle"
    }
    mtn_other.document(date_and_time).set(second_data)
    user22 = mtn_other.document(date_and_time)
    pu = user22.get()
    print(pu.to_dict())
    print("pu")
    mail_doc_ref = mail_collection.document()
    file_path = 'business_api/mtn_maill.txt'  # Replace with your file path

    name = details['first_name']
    volume = data_volume
    date = date_and_time
    reference_t = ref
    receiver_t = receiver

    with open(file_path, 'r') as file:
        html_content = file.read()

    placeholders = {
        '{name}': name,
        '{volume}': volume,
        '{date}': date,
        '{reference}': reference_t,
        '{receiver}': receiver_t
    }

    for placeholder, value in placeholders.items():
        html_content = html_content.replace(placeholder, str(value))

    mail_doc_ref.set({
        'to': details['email'],
        'message': {
            'subject': 'MTN Data',
            'html': html_content,
            'messageId': 'CloudHub GH'
        }
    })
    print("got to redirect")
    return Response(data={'code': '0000', 'message': "Transaction Saved"}, status=status.HTTP_200_OK)


@csrf_exempt
def paystack_webhook(request):
    if request.method == "POST":
        paystack_secret_key = config("PAYSTACK_SECRET_KEY")
        # print(paystack_secret_key)
        payload = json.loads(request.body)

        paystack_signature = request.headers.get("X-Paystack-Signature")

        if not paystack_secret_key or not paystack_signature:
            return HttpResponse(status=400)

        computed_signature = hmac.new(
            paystack_secret_key.encode('utf-8'),
            request.body,
            hashlib.sha512
        ).hexdigest()

        if computed_signature == paystack_signature:
            print("yes")
            print(payload.get('data'))
            r_data = payload.get('data')
            print(r_data.get('metadata'))
            print(payload.get('event'))
            if payload.get('event') == 'charge.success':
                metadata = r_data.get('metadata')
                receiver = metadata.get('receiver')
                bundle_package = metadata.get('bundle_package')
                channel = metadata.get('channel')
                user_id = metadata.get('user_id')
                real_amount = metadata.get('real_amount')
                print(real_amount)
                paid_amount = float(r_data.get('amount')) / 100
                amount = real_amount
                email = r_data.get('email')
                reference = r_data.get('reference')
                date = metadata.get("date")
                time = metadata.get("time")
                date_and_time = metadata.get("date_and_time")
                txn_status = metadata.get("txn_status")

                user_details = get_user_details(user_id)
                if user_details is not None:
                    first_name = user_details['first name']
                    last_name = user_details['last name']
                    email = user_details['email']
                    phone = user_details['phone']
                    first_name = first_name
                else:
                    first_name = ""
                    email = ""

                sms_message = f"Payment of GHS {paid_amount} received at CloudHub. \nReceipt: {reference}\n\nThank You!"
                sms_url = f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UmpEc1JzeFV4cERKTWxUWktqZEs&to={phone}&from=CloudHub GH&sms={sms_message}"
                response = requests.request("GET", url=sms_url)
                print(response.status_code)

                if channel == "ishare":
                    try:
                        send_response = webhook_send_and_save_to_history(user_id=user_id, date_and_time=date_and_time,
                                                                         date=date,
                                                                         time=time,
                                                                         amount=amount, receiver=receiver,
                                                                         reference=reference,
                                                                         paid_at=date_and_time,
                                                                         txn_type="AT Premium Bundle",
                                                                         color_code="Green", data_volume=bundle_package,
                                                                         ishare_balance=0, txn_status=txn_status)
                        print(f"send_response gave us =============================== {send_response}")
                        data = send_response
                        print(f"send_response json gave us =============================== {send_response}")
                        print(data)
                        json_response = data.json()
                        print(json_response)
                        if data.status_code != 200:
                            print("Stopped here")
                            return HttpResponse(status=500)
                        else:
                            print(send_response.status_code)
                            try:
                                batch_id = json_response["batchId"]
                            except KeyError:
                                return HttpResponse(status=200)

                            print(batch_id)

                            if data.status_code == 200:
                                print("enetered into the 200000000000000000000000000000000000000000000000000")
                                sms = f"Hey there\nYour account has been credited with {bundle_package}MB.\nConfirm your new balance using the AT Mobile App"
                                r_sms_url = f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UmpEc1JzeFV4cERKTWxUWktqZEs&to={receiver}&from=CloudHub GH&sms={sms}"
                                response = requests.request("GET", url=r_sms_url)
                                print(response.text)
                                doc_ref = history_collection.document(date_and_time)
                                if doc_ref.get().exists:
                                    doc_ref.update({'done': 'Successful'})
                                else:
                                    print("no entry")
                                mail_doc_ref = mail_collection.document(f"{batch_id}-Mail")
                                file_path = 'business_api/mail.txt'  # Replace with your file path

                                name = first_name
                                volume = bundle_package
                                date = date_and_time
                                reference_t = reference
                                receiver_t = receiver

                                with open(file_path, 'r') as file:
                                    html_content = file.read()

                                placeholders = {
                                    '{name}': name,
                                    '{volume}': volume,
                                    '{date}': date,
                                    '{reference}': reference_t,
                                    '{receiver}': receiver_t
                                }

                                for placeholder, value in placeholders.items():
                                    html_content = html_content.replace(placeholder, str(value))

                                mail_doc_ref.set({
                                    'to': email,
                                    'message': {
                                        'subject': 'AT Flexi Bundle',
                                        'html': html_content,
                                        'messageId': 'CloudHub GH'
                                    }
                                })
                                print("donnee")
                                return HttpResponse(status=200)
                            else:
                                doc_ref = history_collection.document(date_and_time)
                                doc_ref.update({'done': 'Failed'})
                                return HttpResponse(status=200)
                    except:
                        return HttpResponse(status=200)
                elif channel == "mtn_flexi":
                    user_details = get_user_details(user_id)
                    if user_details is not None:
                        first_name = user_details['first name']
                        last_name = user_details['last name']
                        email = user_details['email']
                        phone = user_details['phone']
                    else:
                        first_name = ""
                        last_name = ""
                        email = ""
                        phone = ""
                    details = {
                        'first_name': first_name,
                        'last_name': last_name,
                        'email': email,
                        'user_id': user_id
                    }
                    mtn_response = mtn_flexi_transaction(receiver=receiver, date_and_time=date_and_time, date=date,
                                                         time=time, amount=amount, data_volume=bundle_package,
                                                         channel="MoMo", phone=phone, ref=reference, details=details,
                                                         txn_status=txn_status)
                    print("after mtn responses")
                    if mtn_response.status_code == 200 or mtn_response.data["code"] == "0000":
                        print("mtn donnnneeee")
                        print("yooo")
                        return HttpResponse(status=200)
                    else:
                        return HttpResponse(status=200)
                elif channel == "big-time":
                    if user_details is not None:
                        first_name = user_details['first name']
                        last_name = user_details['last name']
                        email = user_details['email']
                        phone = user_details['phone']
                    else:
                        first_name = ""
                        last_name = ""
                        email = ""
                        phone = ""
                    details = {
                        'first_name': first_name,
                        'last_name': last_name,
                        'email': email,
                        'user_id': user_id
                    }
                    big_time_response = big_time_transaction(receiver=receiver, date_and_time=date_and_time, date=date,
                                                             time=time, amount=amount, data_volume=bundle_package,
                                                             channel="MoMo", phone=phone, ref=reference,
                                                             details=details, txn_status=txn_status)
                    if big_time_response.status_code == 200 or big_time_response.data["code"] == "0000":
                        print("big time donnnneee")
                        return HttpResponse(status=200)
                    else:
                        return HttpResponse(status=200)
                elif channel == "top_up":
                    try:
                        user_details = get_user_details(user_id)
                        if user_details is not None:
                            print(user_details)
                            first_name = user_details['first name']
                            last_name = user_details['last name']
                            email = user_details['email']
                            phone = user_details['phone']
                            try:
                                previous_wallet = user_details['wallet']
                            except KeyError:
                                previous_wallet = 0
                        else:
                            first_name = ""
                            last_name = ""
                            email = ""
                            phone = ""
                            previous_wallet = 0
                        all_data = {
                            'batch_id': "unknown",
                            'buyer': phone,
                            'color_code': "Green",
                            'amount': amount,
                            'data_break_down': amount,
                            'data_volume': bundle_package,
                            'date': date,
                            'date_and_time': date_and_time,
                            'done': "Success",
                            'email': email,
                            'image': user_id,
                            'ishareBalance': 0,
                            'name': f"{first_name} {last_name}",
                            'number': receiver,
                            'paid_at': date_and_time,
                            'reference': reference,
                            'responseCode': 200,
                            'status': txn_status,
                            'time': time,
                            'tranxId': str(tranx_id_generator()),
                            'type': "WALLETTOPUP",
                            'uid': user_id
                        }
                        history_web.collection(email).document(date_and_time).set(all_data)
                        print("f saved")
                        history_collection.document(date_and_time).set(all_data)
                        print(f"ya{history_collection.document(date_and_time).get().to_dict()}")
                        print("f saved")
                        print(f"yo{history_web.collection(email).document(date_and_time).get().to_dict()}")
                        to_be_added = float(amount)
                        print(f"amount to be added: {to_be_added}")
                        new_balance = previous_wallet + to_be_added
                        print(f" new balance: {new_balance}")
                        doc_ref = user_collection.document(user_id)
                        doc_ref.update(
                            {'wallet': new_balance, 'wallet_last_update': date_and_time,
                             'recent_wallet_reference': reference})
                        print("before all data")
                        all_data = {
                            'batch_id': "unknown",
                            'buyer': phone,
                            'color_code': "Green",
                            'amount': amount,
                            'data_break_down': amount,
                            'data_volume': bundle_package,
                            'date': date,
                            'date_and_time': date_and_time,
                            'done': "Success",
                            'email': email,
                            'image': user_id,
                            'ishareBalance': 0,
                            'name': f"{first_name} {last_name}",
                            'number': receiver,
                            'paid_at': date_and_time,
                            'reference': reference,
                            'responseCode': 200,
                            'status': txn_status,
                            'time': time,
                            'tranxId': str(tranx_id_generator()),
                            'type': "WALLETTOPUP",
                            'uid': user_id
                        }
                        print("***********************before saving bla bla****************************")
                        history_web.collection(email).document(date_and_time).set(all_data)
                        print("saved")
                        history_collection.document(date_and_time).set(all_data)
                        print(f"ya{history_collection.document(date_and_time).get().to_dict()}")
                        print("saved")
                        print(f"yo{history_web.collection(email).document(date_and_time).get().to_dict()}")

                        name = f"{first_name} {last_name}"
                        amount = to_be_added
                        file_path = 'business_api/wallet_mail.txt'
                        mail_doc_ref = mail_collection.document()

                        with open(file_path, 'r') as file:
                            html_content = file.read()

                        placeholders = {
                            '{name}': name,
                            '{amount}': amount
                        }

                        for placeholder, value in placeholders.items():
                            html_content = html_content.replace(placeholder, str(value))

                        mail_doc_ref.set({
                            'to': email,
                            'message': {
                                'subject': 'Wallet Topup',
                                'html': html_content,
                                'messageId': 'CloudHub GH'
                            }
                        })
                        print("*****************************before sms*********************************************")
                        sms_message = f"GHS {to_be_added} was deposited in your wallet. Available balance is now GHS {round(new_balance, 2)}"
                        sms_url = f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UmpEc1JzeFV4cERKTWxUWktqZEs&to=0{user_details['phone']}&from=CloudHub GH&sms={sms_message}"
                        response = requests.request("GET", url=sms_url)
                        print(response.status_code)
                        return HttpResponse(status=200)
                    except:
                        return HttpResponse(status=200)
                else:
                    return HttpResponse(status=200)
            else:
                return HttpResponse(status=200)
        else:
            return HttpResponse(status=200)
    else:
        return HttpResponse(status=200)


def hubtel_webhook_send_and_save_to_history(saved_data, user_id, reference, receiver, data_volume, amount,
                                            date_and_time, time):
    user_details = get_user_details(user_id)
    first_name = user_details['first name']
    last_name = user_details['last name']
    email = user_details['email']
    phone = user_details['phone']

    doc_ref = history_web.collection(email).document(reference)
    if doc_ref.get().exists:
        data = doc_ref.get()
        data_dict = data.to_dict()
        batch_id = data_dict["batch_id"]
        if batch_id != "unknown":
            ishare_verification_response = ishare_verification(batch_id)
            if ishare_verification_response is not False:
                code = \
                    ishare_verification_response["flexiIshareTranxStatus"]["flexiIshareTranxStatusResult"][
                        "apiResponse"][
                        "responseCode"]
                ishare_response = \
                    ishare_verification_response["flexiIshareTranxStatus"]["flexiIshareTranxStatusResult"][
                        "ishareApiResponseData"][
                        "apiResponseData"][
                        0][
                        "responseMsg"]
                print(code)
                print(ishare_response)
                if code == '200' or ishare_response == 'Crediting Successful.':
                    return Response(data={"code": "0002"}, status=status.HTTP_200_OK)
                else:
                    pass
        else:
            pass
    else:
        pass
    print("moving on")
    data = saved_data
    history_web.collection(email).document(reference).set(data)

    if history_collection.document(reference).get().exists:
        print("first save")

    ishare_response = send_ishare_bundle(first_name=first_name, last_name=last_name, receiver=receiver,
                                         buyer=phone,
                                         bundle=data_volume,
                                         email=email)
    if ishare_response.status_code == 401:
        return Response(
            data={'status_code': ishare_response.status_code, "message": "Authorization Failed"},
            status=status.HTTP_400_BAD_REQUEST)
    data = ishare_response.json()
    try:
        print("entered the try")
        batch_id = data["batchId"]
        print("batch id")
    except KeyError:
        print("key error")
        return Response(
            data={'status_code': ishare_response.status_code, "message": "Transaction Failed"},
            status=status.HTTP_400_BAD_REQUEST)
    print(data["batchId"])
    status_code = ishare_response.status_code
    if batch_id is None:
        print("batch id was none")
        return Response(data={'status_code': status_code, "message": "Transaction Failed"},
                        status=status.HTTP_400_BAD_REQUEST)
    sleep(10)
    ishare_verification_response = ishare_verification(batch_id)
    if ishare_verification_response is not False:
        code = \
            ishare_verification_response["flexiIshareTranxStatus"]["flexiIshareTranxStatusResult"][
                "apiResponse"][
                "responseCode"]
        ishare_response = \
            ishare_verification_response["flexiIshareTranxStatus"]["flexiIshareTranxStatusResult"][
                "ishareApiResponseData"][
                "apiResponseData"][
                0][
                "responseMsg"]
        print(code)
        print(ishare_response)
        if code == '200' or ishare_response == 'Crediting Successful.':
            sms = f"Your account has been credited with {data_volume}MB."
            r_sms_url = f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UmpEc1JzeFV4cERKTWxUWktqZEs&to={receiver}&from=CloudHub GH&sms={sms}"
            response = requests.request("GET", url=r_sms_url)
            print(response.text)
            doc_ref = history_collection.document(date_and_time)
            doc_ref.update({'done': 'Successful'})
            mail_doc_ref = mail_collection.document(f"{batch_id}-Mail")
            file_path = 'business_api/mail.txt'  # Replace with your file path

            name = user_details["first name"]
            volume = data_volume
            date = date_and_time
            reference_t = reference
            receiver_t = receiver

            with open(file_path, 'r') as file:
                html_content = file.read()

            placeholders = {
                '{name}': name,
                '{volume}': volume,
                '{date}': date,
                '{reference}': reference_t,
                '{receiver}': receiver_t
            }

            for placeholder, value in placeholders.items():
                html_content = html_content.replace(placeholder, str(value))

            mail_doc_ref.set({
                'to': email,
                'message': {
                    'subject': 'AT Flexi Bundle',
                    'html': html_content,
                    'messageId': 'CloudHub GH'
                }
            })

            tot = user_collection.document(user_id)
            print(tot.get().to_dict())
            try:
                print(tot.get().to_dict()['at_total_sales'])
                previous_sale = tot.get().to_dict()['at_total_sales']
                print(f"Previous Sale: {previous_sale}")
                new_sale = float(previous_sale) + float(amount)
                print(new_sale)
                user_collection.document(user_id).update({'at_total_sales': new_sale})
            except:
                user_collection.document(user_id).update({'at_total_sales': amount})

            tat = cashback_collection.document(user_id)
            print(tat.get().to_dict())

            try:
                previous_cashback = tat.get().to_dict()['cashback_wallet']
                print(previous_cashback)
                cashback_balance = (0.5 / 100) * float(amount)
                print(cashback_balance)
                new_cashback = float(previous_cashback) + float(cashback_balance)
                print(new_cashback)
                cashback_collection.document(user_id).update(
                    {'cashback_wallet': new_cashback, 'phone_number': user_details['phone']})
                if models.CashBack.objects.filter(user_id=user_id).exists():
                    user_instance = models.CashBack.objects.filter(user_id=user_id).first()
                    user_instance.amount = new_cashback
                    user_instance.user_number = phone
                    user_instance.save()
                else:
                    new_instance = models.CashBack.objects.create(user_number=phone, user_id=user_id,
                                                                  amount=new_cashback)
                    new_instance.save()
            except TypeError as e:
                print(e)
                cashback_balance = (0.5 / 100) * float(amount)
                print(cashback_balance)
                cashback_collection.document(user_id).set(
                    {'cashback_wallet': cashback_balance, 'phone_number': user_details['phone']})
                if models.CashBack.objects.filter(user_id=user_id).exists():
                    user_instance = models.CashBack.objects.filter(user_id=user_id).first()
                    user_instance.amount = cashback_balance
                    user_instance.user_number = phone
                    user_instance.save()
                else:
                    new_instance = models.CashBack.objects.create(user_number=phone, user_id=user_id,
                                                                  amount=cashback_balance)
                    new_instance.save()
                print(cashback_collection.document(user_id).get().to_dict())
                print("did")

            return Response(data={'status_code': status_code, 'batch_id': batch_id},
                            status=status.HTTP_200_OK)
        else:
            print("some else before history")
            doc_ref = history_collection.document(date_and_time)
            doc_ref.update({'done': 'Failed'})
            return Response(data={'status_code': status_code, "message": "Transaction Failed"},
                            status=status.HTTP_400_BAD_REQUEST)
    else:
        return Response(data={'status_code': status_code, "message": "Could not verify transaction"},
                        status=status.HTTP_400_BAD_REQUEST)
    # json_response = ishare_response.json()
    # print(f"hello:{json_response}")
    # status_code = ishare_response.status_code
    # print(status_code)
    # try:
    #     batch_id = json_response["batch_id"]
    # except KeyError:
    #     batch_id = "unknown"
    # print(batch_id)
    #
    # doc_ref = history_collection.document(reference)
    # if doc_ref.get().exists:
    #     doc_ref.update({'batch_id': batch_id, 'responseCode': status_code})
    #     history_web.collection(email).document(reference).update({'batch_id': batch_id, 'responseCode': status_code})
    # else:
    #     print("didn't find any entry to update")
    # print("firebase saved")
    # # return status_code, batch_id if batch_id else "No batchId", email, first_name
    # return Response(
    #     data=json_response,
    #     status=status.HTTP_200_OK)


def hubtel_mtn_flexi_transaction(saved_data, reference, email, data_volume, date_and_time, receiver, first_name,
                                 user_id, amount, phone):
    history_collection.document(reference).set(saved_data)
    history_web.collection(email).document(reference).set(saved_data)
    user = history_collection.document(reference)
    doc = user.get()
    print(doc.to_dict())
    tranx_id = doc.to_dict()['tranxId']
    second_data = saved_data
    mtn_other.document(reference).set(second_data)
    user22 = mtn_other.document(reference)
    pu = user22.get()
    print(pu.to_dict())
    print("pu")
    mail_doc_ref = mail_collection.document()
    file_path = 'business_api/mtn_maill.txt'  # Replace with your file path

    tot = user_collection.document(user_id)
    print(tot.get().to_dict())
    try:
        print(tot.get().to_dict()['mtn_total_sales'])
        previous_sale = tot.get().to_dict()['mtn_total_sales']
        print(f"Previous Sale: {previous_sale}")
        new_sale = float(previous_sale) + float(amount)
        print(new_sale)
        user_collection.document(user_id).update({'mtn_total_sales': new_sale})
    except:
        user_collection.document(user_id).update({'mtn_total_sales': amount})

    try:
        tot = user_collection.document('9VA0qyq6lXYPZ6Ut867TVcBvF2t1')
        print(tot.get().to_dict()['mtn_total_sales'])
        previous_sale = tot.get().to_dict()['mtn_total_sales']
        print(f"Previous Sale: {previous_sale}")
        new_sale = float(previous_sale) + float(amount)
        print(new_sale)
        user_collection.document('9VA0qyq6lXYPZ6Ut867TVcBvF2t1').update({'mtn_total_sales': new_sale})
    except:
        user_collection.document('9VA0qyq6lXYPZ6Ut867TVcBvF2t1').update({'mtn_total_sales': amount})

    tat = cashback_collection.document(user_id)
    print(tat.get().to_dict())

    tat = cashback_collection.document(user_id)
    print(tat.get().to_dict())

    tat = cashback_collection.document(user_id)
    print(tat.get().to_dict())

    try:
        previous_cashback = tat.get().to_dict()['cashback_wallet']
        print(previous_cashback)
        cashback_balance = (0.5 / 100) * float(amount)
        print(cashback_balance)
        new_cashback = float(previous_cashback) + float(cashback_balance)
        print(new_cashback)
        cashback_collection.document(user_id).update(
            {'cashback_wallet': new_cashback, 'phone_number': phone})
        if models.CashBack.objects.filter(user_id=user_id).exists():
            user_instance = models.CashBack.objects.filter(user_id=user_id).first()
            user_instance.amount = new_cashback
            user_instance.user_number = phone
            user_instance.save()
        else:
            new_instance = models.CashBack.objects.create(user_number=phone, user_id=user_id, amount=new_cashback)
            new_instance.save()
    except TypeError as e:
        print(e)
        cashback_balance = (0.5 / 100) * float(amount)
        print(cashback_balance)
        cashback_collection.document(user_id).set(
            {'cashback_wallet': cashback_balance, 'phone_number': phone})
        if models.CashBack.objects.filter(user_id=user_id).exists():
            user_instance = models.CashBack.objects.filter(user_id=user_id).first()
            user_instance.amount = cashback_balance
            user_instance.user_number = phone
            user_instance.save()
        else:
            new_instance = models.CashBack.objects.create(user_number=phone, user_id=user_id, amount=cashback_balance)
            new_instance.save()
        print(cashback_collection.document(user_id).get().to_dict())
        print("did")
    name = first_name
    volume = data_volume
    date = date_and_time
    reference_t = reference
    receiver_t = receiver

    with open(file_path, 'r') as file:
        html_content = file.read()

    placeholders = {
        '{name}': name,
        '{volume}': volume,
        '{date}': date,
        '{reference}': reference_t,
        '{receiver}': receiver_t
    }

    for placeholder, value in placeholders.items():
        html_content = html_content.replace(placeholder, str(value))

    mail_doc_ref.set({
        'to': email,
        'message': {
            'subject': 'MTN Data',
            'html': html_content,
            'messageId': 'CloudHub GH'
        }
    })
    print("got to redirect")
    return Response(data={'code': '0000', 'message': "Transaction Saved"}, status=status.HTTP_200_OK)


def hubtel_big_time_transaction(saved_data, reference, email, data_volume, date_and_time, receiver, first_name, amount,
                                user_id, phone):
    data = saved_data
    history_collection.document(reference).set(data)
    history_web.collection(email).document(reference).set(saved_data)
    big_time.document(reference).set(saved_data)
    user = history_collection.document(reference)
    doc = user.get()
    print(doc.to_dict())
    tranx_id = doc.to_dict()['tranxId']
    mail_doc_ref = mail_collection.document()
    file_path = 'business_api/mtn_mail.txt'  # Replace with your file path

    name = first_name
    volume = data_volume
    date = date_and_time
    reference_t = reference
    receiver_t = receiver

    with open(file_path, 'r') as file:
        html_content = file.read()

    placeholders = {
        '{name}': name,
        '{volume}': volume,
        '{date}': date,
        '{reference}': reference_t,
        '{receiver}': receiver_t
    }

    tot = user_collection.document(user_id)
    print(tot.get().to_dict())
    try:
        print(tot.get().to_dict()['bt_total_sales'])
        previous_sale = tot.get().to_dict()['bt_total_sales']
        print(f"Previous Sale: {previous_sale}")
        new_sale = float(previous_sale) + float(amount)
        print(new_sale)
        user_collection.document(user_id).update({'bt_total_sales': new_sale})
    except:
        user_collection.document(user_id).update({'bt_total_sales': amount})

    tat = cashback_collection.document(user_id)
    print(tat.get().to_dict())

    try:
        previous_cashback = tat.get().to_dict()['cashback_wallet']
        print(previous_cashback)
        cashback_balance = (0.5 / 100) * float(amount)
        print(cashback_balance)
        new_cashback = float(previous_cashback) + float(cashback_balance)
        print(new_cashback)
        cashback_collection.document(user_id).update(
            {'cashback_wallet': new_cashback, 'phone_number': phone})
        if models.CashBack.objects.filter(user_id=user_id).exists():
            user_instance = models.CashBack.objects.filter(user_id=user_id).first()
            user_instance.amount = new_cashback
            user_instance.user_number = phone
            user_instance.save()
        else:
            new_instance = models.CashBack.objects.create(user_number=phone, user_id=user_id, amount=new_cashback)
            new_instance.save()
    except TypeError as e:
        print(e)
        cashback_balance = (0.5 / 100) * float(amount)
        print(cashback_balance)
        cashback_collection.document(user_id).set(
            {'cashback_wallet': cashback_balance, 'phone_number': phone})
        if models.CashBack.objects.filter(user_id=user_id).exists():
            user_instance = models.CashBack.objects.filter(user_id=user_id).first()
            user_instance.amount = cashback_balance
            user_instance.user_number = phone
            user_instance.save()
        else:
            new_instance = models.CashBack.objects.create(user_number=phone, user_id=user_id, amount=cashback_balance)
            new_instance.save()
        print(cashback_collection.document(user_id).get().to_dict())
        print("did")

    for placeholder, value in placeholders.items():
        html_content = html_content.replace(placeholder, str(value))

    mail_doc_ref.set({
        'to': email,
        'message': {
            'subject': 'Big Time Data',
            'html': html_content,
            'messageId': 'CloudHub GH'
        }
    })
    return Response(data={'code': '0000', 'message': "Transaction Saved"}, status=status.HTTP_200_OK)


@csrf_exempt
def hubtel_webhook(request):
    if request.method == 'POST':
        try:
            payload = request.body.decode('utf-8')
            print("Hubtel payment Info: ", payload)
            json_payload = json.loads(payload)
            print(json_payload)

            data = json_payload.get('Data')
            print(data)
            reference = data.get('ClientReference')
            print(reference)
            txn_status = data.get('Status')
            amount = data.get('Amount')
            print(txn_status, amount)

            # all_data = {
            #     'batch_id': "unknown",
            #     'buyer': phone,
            #     'color_code': "Green",
            #     'amount': amount,
            #     'data_break_down': amount,
            #     'data_volume': bundle_package,
            #     'date': date,
            #     'date_and_time': date_and_time,
            #     'done': "Success",
            #     'email': email,
            #     'image': user_id,
            #     'ishareBalance': 0,
            #     'name': f"{first_name} {last_name}",
            #     'number': receiver,
            #     'paid_at': date_and_time,
            #     'reference': reference,
            #     'responseCode': 200,
            #     'status': txn_status,
            #     'time': time,
            #     'tranxId': str(tranx_id_gen()),
            #     'type': "WALLETTOPUP",
            #     'uid': user_id
            # }

            if txn_status == 'Success':
                print("success")
                collection_saved = history_collection.document(reference).get().to_dict()
                receiver = collection_saved['number']
                bundle_volume = collection_saved['data_volume']
                name = collection_saved['name']
                email = collection_saved['email']
                phone_number = collection_saved['buyer']
                date_and_time = collection_saved['date_and_time']
                time = collection_saved["time"]
                txn_type = collection_saved['type']
                user_id = collection_saved['uid']
                that_amount = collection_saved['amount']
                print(receiver, bundle_volume, name, email, phone_number)

                doc_ref = history_collection.document(reference)

                sms_message = f"Payment of GHS {amount} received at CloudHub. \nReceipt: {reference}\n\nThank You!"
                sms_url = f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UmpEc1JzeFV4cERKTWxUWktqZEs&to={phone_number}&from=CloudHub GH&sms={sms_message}"
                response = requests.request("GET", url=sms_url)
                print(response.status_code)

                if txn_type == "AT Premium Internet":
                    doc_ref.update(
                        {'ishareBalance': "Paid", 'status': "Delivered", "tranxId": str(tranx_id_generator())})
                    user_details = get_user_details(user_id)
                    if user_details is not None:
                        first_name = user_details['first name']
                        last_name = user_details['last name']
                        email = user_details['email']
                        phone = user_details['phone']
                    else:
                        first_name = ""
                        last_name = ""
                        email = ""
                        phone = ""
                    collection_saved = history_collection.document(reference).get().to_dict()
                    send_response = hubtel_webhook_send_and_save_to_history(collection_saved, user_id, reference,
                                                                            receiver, bundle_volume, amount,
                                                                            date_and_time, time)
                    # saved_data, user_id, reference, receiver, data_volume
                    print(send_response)
                    return HttpResponse(status=send_response.status_code)

                elif txn_type == "MTN Master Internet":
                    doc_ref.update(
                        {'ishareBalance': "Paid", 'status': "Undelivered", "tranxId": str(tranx_id_generator())})

                    new_mtn_txn = models.MTNTransaction.objects.create(
                        user_id=user_id,
                        amount=amount,
                        bundle_volume=bundle_volume,
                        number=receiver,
                        firebase_date=date_and_time
                    )
                    new_mtn_txn.save()
                    user_details = get_user_details(user_id)
                    if user_details is not None:
                        first_name = user_details['first name']
                        last_name = user_details['last name']
                        email = user_details['email']
                        phone = user_details['phone']
                    else:
                        first_name = ""
                        last_name = ""
                        email = ""
                        phone = ""
                    details = {
                        'first_name': first_name,
                        'last_name': last_name,
                        'email': email,
                        'user_id': user_id
                    }
                    collection_saved = history_collection.document(reference).get().to_dict()
                    mtn_response = hubtel_mtn_flexi_transaction(collection_saved, reference, email, bundle_volume,
                                                                date_and_time, receiver, first_name, user_id,
                                                                that_amount, phone)
                    print(mtn_response)
                    # new_mtn_txn = models.MTNTransaction.objects.create(
                    #     user_id=user_id,
                    #     amount=amount,
                    #     bundle_volume=bundle_volume,
                    #     number=receiver,
                    #     firebase_date=date_and_time
                    # )
                    # new_mtn_txn.save()
                    # saved_data, reference, email, data_volume, date_and_time, receiver, first_name
                    print("after mtn responses")
                    if mtn_response.status_code == 200 or mtn_response.data["code"] == "0000":
                        print("mtn donnnneeee")
                        print("yooo")
                        return HttpResponse(status=200)
                    else:
                        return JsonResponse({'message': "Success"}, status=200)
                elif txn_type == "AT Big Time Internet":
                    doc_ref.update(
                        {'ishareBalance': "Paid", 'status': "Undelivered", "tranxId": str(tranx_id_generator())})
                    user_details = get_user_details(user_id)
                    if user_details is not None:
                        first_name = user_details['first name']
                        last_name = user_details['last name']
                        email = user_details['email']
                        phone = user_details['phone']
                    else:
                        first_name = ""
                        last_name = ""
                        email = ""
                        phone = ""
                    details = {
                        'first_name': first_name,
                        'last_name': last_name,
                        'email': email,
                        'user_id': user_id
                    }
                    collection_saved = history_collection.document(reference).get().to_dict()
                    big_time_response = hubtel_big_time_transaction(collection_saved, reference, email, bundle_volume,
                                                                    date_and_time, receiver, first_name, amount,
                                                                    user_id, phone)
                    # saved_data, reference, email, data_volume, date_and_time, receiver, first_name
                    if big_time_response.status_code == 200 or big_time_response.data["code"] == "0000":
                        print("big time donnnneee")
                        return JsonResponse({'message': "Success"}, status=200)
                    else:
                        return HttpResponse(status=500)
                elif txn_type == "CloudHub E-Wallet":
                    doc_ref.update(
                        {'ishareBalance': "Paid", 'status': "Credited", "tranxId": str(tranx_id_generator())})
                    user_details = get_user_details(user_id)
                    collection_saved = history_collection.document(reference).get().to_dict()
                    if user_details is not None:
                        print(user_details)
                        first_name = user_details['first name']
                        last_name = user_details['last name']
                        email = user_details['email']
                        phone = user_details['phone']
                        try:
                            previous_wallet = user_details['wallet']
                        except KeyError:
                            previous_wallet = 0
                    else:
                        first_name = ""
                        last_name = ""
                        email = ""
                        phone = ""
                        previous_wallet = 0
                    all_data = collection_saved
                    history_web.collection(email).document(date_and_time).set(all_data)
                    print("f saved")
                    history_collection.document(date_and_time).set(all_data)
                    print(f"ya{history_collection.document(date_and_time).get().to_dict()}")
                    print("f saved")
                    print(f"yo{history_web.collection(email).document(date_and_time).get().to_dict()}")
                    to_be_added = float(that_amount)
                    print(f"amount to be added: {to_be_added}")
                    new_balance = previous_wallet + to_be_added
                    print(f" new balance: {new_balance}")
                    doc_ref = user_collection.document(user_id)
                    doc_ref.update(
                        {'wallet': new_balance, 'wallet_last_update': date_and_time,
                         'recent_wallet_reference': reference})
                    print(doc_ref.get().to_dict())
                    print("before all data")
                    all_data = collection_saved
                    history_web.collection(email).document(date_and_time).set(all_data)
                    print("saved")
                    print(f"ya{history_collection.document(date_and_time).get().to_dict()}")
                    print("saved")
                    print(f"yo{history_web.collection(email).document(date_and_time).get().to_dict()}")

                    name = f"{first_name} {last_name}"
                    amount = to_be_added
                    file_path = 'business_api/wallet_mail.txt'
                    mail_doc_ref = mail_collection.document()

                    with open(file_path, 'r') as file:
                        html_content = file.read()

                    placeholders = {
                        '{name}': name,
                        '{amount}': amount
                    }

                    for placeholder, value in placeholders.items():
                        html_content = html_content.replace(placeholder, str(value))

                    mail_doc_ref.set({
                        'to': email,
                        'message': {
                            'subject': 'Wallet Topup',
                            'html': html_content,
                            'messageId': 'CloudHub GH'
                        }
                    })

                    sms_message = f"GHS {to_be_added} was deposited in your wallet. Available balance is now GHS {new_balance}"
                    sms_url = f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UmpEc1JzeFV4cERKTWxUWktqZEs&to=0{user_details['phone']}&from=CloudHub GH&sms={sms_message}"
                    response = requests.request("GET", url=sms_url)
                    print(response.status_code)
                    return JsonResponse({'message': "Success"}, status=200)
                elif txn_type == "ISHARE_TOPUP":
                    doc_ref.update(
                        {'ishareBalance': "Paid", 'status': "Credited", "tranxId": str(tranx_id_generator())})
                    user_details = get_user_details(user_id)
                    collection_saved = history_collection.document(reference).get().to_dict()
                    if user_details is not None:
                        print(user_details)
                        first_name = user_details['first name']
                        last_name = user_details['last name']
                        email = user_details['email']
                        phone = user_details['phone']
                        try:
                            previous_wallet = user_details['at_balance']
                        except KeyError:
                            previous_wallet = 0
                    else:
                        first_name = ""
                        last_name = ""
                        email = ""
                        phone = ""
                        previous_wallet = 0
                    all_data = collection_saved
                    history_web.collection(email).document(date_and_time).set(all_data)
                    print("f saved")
                    history_collection.document(date_and_time).set(all_data)
                    print(f"ya{history_collection.document(date_and_time).get().to_dict()}")
                    print("f saved")
                    print(f"yo{history_web.collection(email).document(date_and_time).get().to_dict()}")
                    to_be_added = float(amount)
                    print(f"amount to be added: {to_be_added}")
                    new_balance = previous_wallet + to_be_added
                    print(f" new balance: {new_balance}")
                    doc_ref = user_collection.document(user_id)
                    doc_ref.update(
                        {'at_balance': new_balance})
                    print(doc_ref.get().to_dict())
                    print("before all data")
                    all_data = collection_saved
                    history_web.collection(email).document(date_and_time).set(all_data)
                    print("saved")
                    print(f"ya{history_collection.document(date_and_time).get().to_dict()}")
                    print("saved")
                    print(f"yo{history_web.collection(email).document(date_and_time).get().to_dict()}")

                    name = f"{first_name} {last_name}"
                    amount = to_be_added
                    file_path = 'business_api/wallet_mail.txt'
                    mail_doc_ref = mail_collection.document()

                    with open(file_path, 'r') as file:
                        html_content = file.read()

                    placeholders = {
                        '{name}': name,
                        '{amount}': amount
                    }

                    for placeholder, value in placeholders.items():
                        html_content = html_content.replace(placeholder, str(value))

                    mail_doc_ref.set({
                        'to': email,
                        'message': {
                            'subject': 'Wallet Topup',
                            'html': html_content,
                            'messageId': 'CloudHub GH'
                        }
                    })

                    sms_message = f"GHS {to_be_added} was deposited in your AT Wallet. Available AT Wallet balance is now GHS {new_balance}"
                    sms_url = f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UmpEc1JzeFV4cERKTWxUWktqZEs&to=0{user_details['phone']}&from=CloudHub GH&sms={sms_message}"
                    response = requests.request("GET", url=sms_url)
                    print(response.status_code)
                    return JsonResponse({'message': "Success"}, status=200)
                elif txn_type == "MTN_AIRTIME":
                    doc_ref.update(
                        {'ishareBalance': "Paid", 'status': "Credited", "tranxId": str(tranx_id_generator())})
                    user_details = get_user_details(user_id)
                    if user_details is not None:
                        print(user_details)
                        first_name = user_details['first name']
                        last_name = user_details['last name']
                        email = user_details['email']
                        phone = user_details['phone']
                    else:
                        first_name = ""
                        last_name = ""
                        email = ""
                        phone = ""
                    url = "https://cs.hubtel.com/commissionservices/2019604/fdd76c884e614b1c8f669a3207b09a98"
                    print(url)

                    print(f"receiver: {receiver}")

                    payload = json.dumps({
                        "Destination": receiver,
                        "Amount": that_amount,
                        "CallbackURL": "https://merchant.cloudhubgh.com/hubtel_webhook",
                        "ClientReference": reference
                    })
                    headers = {
                        'Authorization': config("HUBTEL_API_KEY"),
                        'Content-Type': 'application/json'
                    }

                    response = requests.request("POST", url, headers=headers, data=payload)

                    print(response.text)

                    if response.status_code == 200:
                        doc_ref = history_collection.document(date_and_time)
                        if doc_ref.get().exists:
                            doc_ref.update({'done': 'Successful'})
                        else:
                            print("no entry")
                        print("worked well")
                        return Response(
                            {"status": response.status_code, 'message': f'Transaction Completed Successfully'},
                        )
                    else:
                        print("not 200 error")
                        return Response({"status": response.status_code, 'message': f'Something went wrong'},
                                        status=400)
                elif txn_type == "VODA_AIRTIME":
                    doc_ref.update(
                        {'ishareBalance': "Paid", 'status': "Credited", "tranxId": str(tranx_id_generator())})
                    user_details = get_user_details(user_id)
                    if user_details is not None:
                        print(user_details)
                        first_name = user_details['first name']
                        last_name = user_details['last name']
                        email = user_details['email']
                        phone = user_details['phone']
                    else:
                        first_name = ""
                        last_name = ""
                        email = ""
                        phone = ""
                    url = "https://cs.hubtel.com/commissionservices/2019604/f4be83ad74c742e185224fdae1304800"
                    print(url)

                    print(f"receiver: {receiver}")

                    payload = json.dumps({
                        "Destination": receiver,
                        "Amount": that_amount,
                        "CallbackURL": "https://merchant.cloudhubgh.com/hubtel_webhook",
                        "ClientReference": reference
                    })
                    headers = {
                        'Authorization': config("HUBTEL_API_KEY"),
                        'Content-Type': 'application/json'
                    }

                    response = requests.request("POST", url, headers=headers, data=payload)

                    print(response.text)

                    if response.status_code == 200:
                        doc_ref = history_collection.document(date_and_time)
                        if doc_ref.get().exists:
                            doc_ref.update({'done': 'Successful'})
                        else:
                            print("no entry")
                        print("worked well")
                        return Response(
                            {"status": response.status_code, 'message': f'Transaction Completed Successfully'},
                        )
                    else:
                        print("not 200 error")
                        return Response({"status": response.status_code, 'message': f'Something went wrong'},
                                        status=400)
                elif txn_type == "AT_AIRTIME":
                    doc_ref.update(
                        {'ishareBalance': "Paid", 'status': "Credited", "tranxId": str(tranx_id_generator())})
                    user_details = get_user_details(user_id)
                    if user_details is not None:
                        print(user_details)
                        first_name = user_details['first name']
                        last_name = user_details['last name']
                        email = user_details['email']
                        phone = user_details['phone']
                    else:
                        first_name = ""
                        last_name = ""
                        email = ""
                        phone = ""
                    url = "https://cs.hubtel.com/commissionservices/2019604/dae2142eb5a14c298eace60240c09e4b"
                    print(url)

                    print(f"receiver: {receiver}")

                    payload = json.dumps({
                        "Destination": receiver,
                        "Amount": that_amount,
                        "CallbackURL": "https://merchant.cloudhubgh.com/hubtel_webhook",
                        "ClientReference": reference
                    })
                    headers = {
                        'Authorization': config("HUBTEL_API_KEY"),
                        'Content-Type': 'application/json'
                    }

                    response = requests.request("POST", url, headers=headers, data=payload)

                    print(response.text)

                    if response.status_code == 200:
                        doc_ref = history_collection.document(date_and_time)
                        if doc_ref.get().exists:
                            doc_ref.update({'done': 'Successful'})
                        else:
                            print("no entry")
                        print("worked well")
                        return Response(
                            {"status": response.status_code, 'message': f'Transaction Completed Successfully'},
                        )
                    else:
                        print("not 200 error")
                        return Response({"status": response.status_code, 'message': f'Something went wrong'},
                                        status=400)
                elif txn_type == "GLO_AIRTIME":
                    doc_ref.update(
                        {'ishareBalance': "Paid", 'status': "Credited", "tranxId": str(tranx_id_generator())})
                    user_details = get_user_details(user_id)
                    if user_details is not None:
                        print(user_details)
                        first_name = user_details['first name']
                        last_name = user_details['last name']
                        email = user_details['email']
                        phone = user_details['phone']
                    else:
                        first_name = ""
                        last_name = ""
                        email = ""
                        phone = ""
                    url = "https://cs.hubtel.com/commissionservices/2019604/47d88e88f50f47468a34a14ac73e8ab5"
                    print(url)

                    print(f"receiver: {receiver}")

                    payload = json.dumps({
                        "Destination": receiver,
                        "Amount": that_amount,
                        "CallbackURL": "https://merchant.cloudhubgh.com/hubtel_webhook",
                        "ClientReference": reference
                    })
                    headers = {
                        'Authorization': config("HUBTEL_API_KEY"),
                        'Content-Type': 'application/json'
                    }

                    response = requests.request("POST", url, headers=headers, data=payload)

                    print(response.text)

                    if response.status_code == 200:
                        doc_ref = history_collection.document(date_and_time)
                        if doc_ref.get().exists:
                            doc_ref.update({'done': 'Successful'})
                        else:
                            print("no entry")
                        print("worked well")
                        return Response(
                            {"status": response.status_code, 'message': f'Transaction Completed Successfully'},
                        )
                    else:
                        print("not 200 error")
                        return Response({"status": response.status_code, 'message': f'Something went wrong'},
                                        status=400)
                else:
                    print("no type found")
                    return JsonResponse({'message': "No Type Found"}, status=500)
            else:
                doc_ref = history_collection.document(reference)
                doc_ref.update({'status': "Failed"})
                return JsonResponse({'message': 'Transaction Failed'}, status=200)
        except Exception as e:
            print("Error Processing hubtel webhook:", str(e))
            return JsonResponse({'status': 'error'}, status=500)
    else:
        print("not post")
        return JsonResponse({'message': 'Not Found'}, status=404)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@authentication_classes([BearerTokenAuthentication])
def initiate_at_airtime(request):
    authorization_header = request.headers.get('Authorization')
    if authorization_header:
        auth_type, token = authorization_header.split(' ')
        if auth_type == 'Bearer':
            try:
                token_obj = Token.objects.get(key=token)
                user = token_obj.user
                user_id = user.user_id
                print(user_id)

                print("hiiiii")

                receiver = request.data.get('receiver')
                reference = request.data.get('reference')
                amount = request.data.get('amount')
                print(amount)

                print("yo")

                if not receiver or not reference or not amount:
                    return Response({'message': 'Body parameters not valid. Check and try again.'},
                                    status=status.HTTP_400_BAD_REQUEST)

                print("got here")
                user_details = get_user_details(user_id)
                print(user_details['first name'])

                date = datetime.datetime.now().strftime("%a, %b %d, %Y")
                time = datetime.datetime.now().strftime("%I:%M:%S %p")
                date_and_time = datetime.datetime.now().isoformat()

                if "wallet" == "wallet":
                    try:
                        enough_balance = check_user_balance_against_price(user_id, amount)
                    except:
                        return Response(
                            {'code': '0001', 'message': f'User ID does not exist: User ID provided: {user_id}.'},
                            status=status.HTTP_400_BAD_REQUEST)
                else:
                    enough_balance = True
                    print("not wallet")
                print(enough_balance)
                if enough_balance:
                    user_details = get_user_details(user_id)
                    email = user_details['email']
                    print(enough_balance)
                    # hist = history_web.collection(email).document(date_and_time)
                    # doc = hist.get()
                    # if doc.exists:
                    #     return redirect(f"https://{callback_url}")
                    # else:
                    #     print("no record found")
                    if "wallet" == "wallet":
                        user = get_user_details(user_id)
                        if user is None:
                            return None
                        previous_user_wallet = user['wallet']
                        print(f"previous wallet: {previous_user_wallet}")
                        new_balance = float(previous_user_wallet) - float(amount)
                        print(f"new_balance:{new_balance}")
                        doc_ref = user_collection.document(user_id)
                        doc_ref.update({'wallet': new_balance})
                        user = get_user_details(user_id)
                        new_user_wallet = user['wallet']
                        print(f"new_user_wallet: {new_user_wallet}")
                        if new_user_wallet == previous_user_wallet:
                            user = get_user_details(user_id)
                            if user is None:
                                return None
                            previous_user_wallet = user['wallet']
                            print(f"previous wallet: {previous_user_wallet}")
                            new_balance = float(previous_user_wallet) - float(amount)
                            print(f"new_balance:{new_balance}")
                            doc_ref = user_collection.document(user_id)
                            doc_ref.update({'wallet': new_balance})
                            user = get_user_details(user_id)
                            new_user_wallet = user['wallet']
                            print(f"new_user_wallet: {new_user_wallet}")
                        else:
                            print("it's fine")
                else:
                    return Response({"code": 400, "message": "Insufficient Balance"},
                                    status=status.HTTP_400_BAD_REQUEST)
                if user_details is not None:
                    print("yes")
                    first_name = user_details['first name']
                    print(first_name)
                    last_name = user_details['last name']
                    print(last_name)
                    email = user_details['email']
                    phone = user_details['phone']
                else:
                    first_name = ""
                    last_name = ""
                    email = ""
                    phone = ""
                details = {
                    'first_name': first_name,
                    'last_name': last_name,
                    'email': email,
                    'user_id': user_id
                }

                url = "https://cs.hubtel.com/commissionservices/2019604/dae2142eb5a14c298eace60240c09e4b"

                payload = json.dumps({
                    "Destination": str(receiver),
                    "Amount": amount,
                    "CallbackURL": "https://merchant.cloudhubgh.com/hubtel_webhook",
                    "ClientReference": reference
                })
                headers = {
                    'Authorization': config("HUBTEL_API_KEY"),
                    'Content-Type': 'application/json'
                }

                response = requests.request("POST", url, headers=headers, data=payload)

                print(response.text)

                if response.status_code == 200:
                    data = {
                        'batch_id': "unknown",
                        'buyer': phone,
                        'color_code': "Green",
                        'amount': amount,
                        'data_break_down': "",
                        'data_volume': "",
                        'date': date,
                        'date_and_time': date_and_time,
                        'done': "unknown",
                        'email': email,
                        'image': user_id,
                        'ishareBalance': 0,
                        'name': f"{first_name} {last_name}",
                        'number': receiver,
                        'paid_at': date_and_time,
                        'reference': reference,
                        'responseCode': "0",
                        'status': "Delivered",
                        'time': time,
                        'tranxId': str(tranx_id_generator()),
                        'type': "AT Airtime",
                        'uid': user_id,
                        'bal': user["wallet"]
                    }
                    history_collection.document(date_and_time).set(data)
                    history_web.collection(email).document(date_and_time).set(data)
                    print("worked well")
                    return Response({"status": response.status_code, 'message': f'Transaction Completed Successfully'},
                                    )
                else:
                    print("not 200 error")
                    return Response({"status": response.status_code, 'message': f'Something went wrong'},
                                    status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"status": '400', 'message': f'Something went wrong: {e}'},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'error': 'Invalid Header Provided.'}, status=status.HTTP_401_UNAUTHORIZED)
    else:
        return Response({'error': 'Invalid Header Provided.'}, status=status.HTTP_401_UNAUTHORIZED)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@authentication_classes([BearerTokenAuthentication])
def admin_initiate_at_airtime(request):
    authorization_header = request.headers.get('Authorization')
    if authorization_header:
        auth_type, token = authorization_header.split(' ')
        if auth_type == 'Bearer':
            try:
                token_obj = Token.objects.get(key=token)
                token_key = token_obj.key
                if token_key != config("TOKEN_KEY"):
                    return Response({'message': 'Authorisation Failed.'},
                                    status=status.HTTP_400_BAD_REQUEST)

                receiver = request.data.get('receiver')
                reference = request.data.get('reference')
                amount = request.data.get('amount')
                user_id = request.data.get('user_id')
                print(amount)

                print("yo")

                if not receiver or not reference or not amount:
                    return Response({'message': 'Body parameters not valid. Check and try again.'},
                                    status=status.HTTP_400_BAD_REQUEST)

                print("got here")
                user_details = get_user_details(user_id)
                print(user_details['first name'])

                date = datetime.datetime.now().strftime("%a, %b %d, %Y")
                time = datetime.datetime.now().strftime("%I:%M:%S %p")
                date_and_time = datetime.datetime.now().isoformat()

                if "wallet" == "wallet":
                    try:
                        enough_balance = check_user_balance_against_price(user_id, amount)
                    except:
                        return Response(
                            {'code': '0001', 'message': f'User ID does not exist: User ID provided: {user_id}.'},
                            status=status.HTTP_400_BAD_REQUEST)
                else:
                    enough_balance = True
                    print("not wallet")
                print(enough_balance)
                if enough_balance:
                    user_details = get_user_details(user_id)
                    email = user_details['email']
                    print(enough_balance)
                    # hist = history_web.collection(email).document(date_and_time)
                    # doc = hist.get()
                    # if doc.exists:
                    #     return redirect(f"https://{callback_url}")
                    # else:
                    #     print("no record found")
                    if "wallet" == "wallet":
                        user = get_user_details(user_id)
                        if user is None:
                            return None
                        previous_user_wallet = user['wallet']
                        print(f"previous wallet: {previous_user_wallet}")
                        new_balance = float(previous_user_wallet) - float(amount)
                        print(f"new_balance:{new_balance}")
                        doc_ref = user_collection.document(user_id)
                        doc_ref.update({'wallet': new_balance})
                        user = get_user_details(user_id)
                        new_user_wallet = user['wallet']
                        print(f"new_user_wallet: {new_user_wallet}")
                        if new_user_wallet == previous_user_wallet:
                            user = get_user_details(user_id)
                            if user is None:
                                return None
                            previous_user_wallet = user['wallet']
                            print(f"previous wallet: {previous_user_wallet}")
                            new_balance = float(previous_user_wallet) - float(amount)
                            print(f"new_balance:{new_balance}")
                            doc_ref = user_collection.document(user_id)
                            doc_ref.update({'wallet': new_balance})
                            user = get_user_details(user_id)
                            new_user_wallet = user['wallet']
                            print(f"new_user_wallet: {new_user_wallet}")
                        else:
                            print("it's fine")
                else:
                    return Response({"code": 400, "message": "Insufficient Balance"},
                                    status=status.HTTP_400_BAD_REQUEST)
                if user_details is not None:
                    print("yes")
                    first_name = user_details['first name']
                    print(first_name)
                    last_name = user_details['last name']
                    print(last_name)
                    email = user_details['email']
                    phone = user_details['phone']
                else:
                    first_name = ""
                    last_name = ""
                    email = ""
                    phone = ""
                details = {
                    'first_name': first_name,
                    'last_name': last_name,
                    'email': email,
                    'user_id': user_id
                }

                url = "https://cs.hubtel.com/commissionservices/2019604/dae2142eb5a14c298eace60240c09e4b"

                payload = json.dumps({
                    "Destination": str(receiver),
                    "Amount": amount,
                    "CallbackURL": "https://merchant.cloudhubgh.com/hubtel_webhook",
                    "ClientReference": reference
                })
                headers = {
                    'Authorization': config("HUBTEL_API_KEY"),
                    'Content-Type': 'application/json'
                }

                response = requests.request("POST", url, headers=headers, data=payload)

                print(response.text)

                if response.status_code == 200:
                    data = {
                        'batch_id': "unknown",
                        'buyer': phone,
                        'color_code': "Green",
                        'amount': amount,
                        'data_break_down': "",
                        'data_volume': "",
                        'date': date,
                        'date_and_time': date_and_time,
                        'done': "unknown",
                        'email': email,
                        'image': user_id,
                        'ishareBalance': 0,
                        'name': f"{first_name} {last_name}",
                        'number': receiver,
                        'paid_at': date_and_time,
                        'reference': reference,
                        'responseCode': "0",
                        'status': "Delivered",
                        'time': time,
                        'tranxId': str(tranx_id_generator()),
                        'type': "AT Airtime",
                        'uid': user_id,
                        'bal': user_details["wallet"]
                    }
                    history_collection.document(date_and_time).set(data)
                    history_web.collection(email).document(date_and_time).set(data)
                    print("worked well")
                    return Response({"status": response.status_code, 'message': f'Transaction Completed Successfully'},
                                    )
                else:
                    print("not 200 error")
                    return Response({"status": response.status_code, 'message': f'Something went wrong'},
                                    status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"status": '400', 'message': f'Something went wrong: {e}'},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'error': 'Invalid Header Provided.'}, status=status.HTTP_401_UNAUTHORIZED)
    else:
        return Response({'error': 'Invalid Header Provided.'}, status=status.HTTP_401_UNAUTHORIZED)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@authentication_classes([BearerTokenAuthentication])
def initiate_mtn_airtime(request):
    authorization_header = request.headers.get('Authorization')
    if authorization_header:
        auth_type, token = authorization_header.split(' ')
        if auth_type == 'Bearer':
            try:
                token_obj = Token.objects.get(key=token)
                user = token_obj.user
                user_id = user.user_id
                print(user_id)

                print("hiiiii")

                receiver = request.data.get('receiver')
                reference = request.data.get('reference')
                amount = request.data.get('amount')
                print(amount)

                print("yo")

                if not receiver or not reference or not amount:
                    return Response({'message': 'Body parameters not valid. Check and try again.'},
                                    status=status.HTTP_400_BAD_REQUEST)

                print("got here")
                user_details = get_user_details(user_id)
                print(user_details['first name'])

                date = datetime.datetime.now().strftime("%a, %b %d, %Y")
                time = datetime.datetime.now().strftime("%I:%M:%S %p")
                date_and_time = datetime.datetime.now().isoformat()

                if "wallet" == "wallet":
                    try:
                        enough_balance = check_user_balance_against_price(user_id, amount)
                    except:
                        return Response(
                            {'code': '0001', 'message': f'User ID does not exist: User ID provided: {user_id}.'},
                            status=status.HTTP_400_BAD_REQUEST)
                else:
                    enough_balance = True
                    print("not wallet")
                print(enough_balance)
                if enough_balance:
                    user_details = get_user_details(user_id)
                    email = user_details['email']
                    print(enough_balance)
                    # hist = history_web.collection(email).document(date_and_time)
                    # doc = hist.get()
                    # if doc.exists:
                    #     return redirect(f"https://{callback_url}")
                    # else:
                    #     print("no record found")
                    if "wallet" == "wallet":
                        user = get_user_details(user_id)
                        if user is None:
                            return None
                        previous_user_wallet = user['wallet']
                        print(f"previous wallet: {previous_user_wallet}")
                        new_balance = float(previous_user_wallet) - float(amount)
                        print(f"new_balance:{new_balance}")
                        doc_ref = user_collection.document(user_id)
                        doc_ref.update({'wallet': new_balance})
                        user = get_user_details(user_id)
                        new_user_wallet = user['wallet']
                        print(f"new_user_wallet: {new_user_wallet}")
                        if new_user_wallet == previous_user_wallet:
                            user = get_user_details(user_id)
                            if user is None:
                                return None
                            previous_user_wallet = user['wallet']
                            print(f"previous wallet: {previous_user_wallet}")
                            new_balance = float(previous_user_wallet) - float(amount)
                            print(f"new_balance:{new_balance}")
                            doc_ref = user_collection.document(user_id)
                            doc_ref.update({'wallet': new_balance})
                            user = get_user_details(user_id)
                            new_user_wallet = user['wallet']
                            print(f"new_user_wallet: {new_user_wallet}")
                        else:
                            print("it's fine")

                if user_details is not None:
                    print("yes")
                    first_name = user_details['first name']
                    print(first_name)
                    last_name = user_details['last name']
                    print(last_name)
                    email = user_details['email']
                    phone = user_details['phone']
                else:
                    first_name = ""
                    last_name = ""
                    email = ""
                    phone = ""
                details = {
                    'first_name': first_name,
                    'last_name': last_name,
                    'email': email,
                    'user_id': user_id
                }

                url = "https://cs.hubtel.com/commissionservices/2019604/fdd76c884e614b1c8f669a3207b09a98"

                payload = json.dumps({
                    "Destination": str(receiver),
                    "Amount": amount,
                    "CallbackURL": "https://merchant.cloudhubgh.com/hubtel_webhook",
                    "ClientReference": reference
                })
                headers = {
                    'Authorization': config("HUBTEL_API_KEY"),
                    'Content-Type': 'application/json'
                }

                response = requests.request("POST", url, headers=headers, data=payload)

                print(response.text)

                if response.status_code == 200:
                    data = {
                        'batch_id': "unknown",
                        'buyer': phone,
                        'color_code': "Green",
                        'amount': amount,
                        'data_break_down': "",
                        'data_volume': "",
                        'date': date,
                        'date_and_time': date_and_time,
                        'done': "unknown",
                        'email': email,
                        'image': user_id,
                        'ishareBalance': 0,
                        'name': f"{first_name} {last_name}",
                        'number': receiver,
                        'paid_at': date_and_time,
                        'reference': reference,
                        'responseCode': "0",
                        'status': "Delivered",
                        'time': time,
                        'tranxId': str(tranx_id_generator()),
                        'type': "MTN Airtime",
                        'uid': user_id,
                        'bal': user["wallet"]
                    }
                    history_collection.document(date_and_time).set(data)
                    history_web.collection(email).document(date_and_time).set(data)
                    print("worked well")
                    return Response({"status": response.status_code, 'message': f'Transaction Completed Successfully'},
                                    )
                else:
                    print("not 200 error")
                    return Response({"status": response.status_code, 'message': f'Something went wrong'},
                                    status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"status": '400', 'message': f'Something went wrong: {e}'},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'error': 'Invalid Header Provided.'}, status=status.HTTP_401_UNAUTHORIZED)
    else:
        return Response({'error': 'Invalid Header Provided.'}, status=status.HTTP_401_UNAUTHORIZED)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@authentication_classes([BearerTokenAuthentication])
def admin_initiate_mtn_airtime(request):
    authorization_header = request.headers.get('Authorization')
    if authorization_header:
        auth_type, token = authorization_header.split(' ')
        if auth_type == 'Bearer':
            try:
                token_obj = Token.objects.get(key=token)
                token_key = token_obj.key
                if token_key != config("TOKEN_KEY"):
                    return Response({'message': 'Authorisation Failed.'},
                                    status=status.HTTP_400_BAD_REQUEST)

                receiver = request.data.get('receiver')
                reference = request.data.get('reference')
                amount = request.data.get('amount')
                user_id = request.data.get('user_id')
                print(amount)

                print("yo")

                if not receiver or not reference or not amount:
                    return Response({'message': 'Body parameters not valid. Check and try again.'},
                                    status=status.HTTP_400_BAD_REQUEST)

                print("got here")
                user_details = get_user_details(user_id)
                print(user_details['first name'])

                date = datetime.datetime.now().strftime("%a, %b %d, %Y")
                time = datetime.datetime.now().strftime("%I:%M:%S %p")
                date_and_time = datetime.datetime.now().isoformat()

                if "wallet" == "wallet":
                    try:
                        enough_balance = check_user_balance_against_price(user_id, amount)
                    except:
                        return Response(
                            {'code': '0001', 'message': f'User ID does not exist: User ID provided: {user_id}.'},
                            status=status.HTTP_400_BAD_REQUEST)
                else:
                    enough_balance = True
                    print("not wallet")
                print(enough_balance)
                if enough_balance:
                    user_details = get_user_details(user_id)
                    email = user_details['email']
                    print(enough_balance)
                    # hist = history_web.collection(email).document(date_and_time)
                    # doc = hist.get()
                    # if doc.exists:
                    #     return redirect(f"https://{callback_url}")
                    # else:
                    #     print("no record found")
                    if "wallet" == "wallet":
                        user = get_user_details(user_id)
                        if user is None:
                            return None
                        previous_user_wallet = user['wallet']
                        print(f"previous wallet: {previous_user_wallet}")
                        new_balance = float(previous_user_wallet) - float(amount)
                        print(f"new_balance:{new_balance}")
                        doc_ref = user_collection.document(user_id)
                        doc_ref.update({'wallet': new_balance})
                        user = get_user_details(user_id)
                        new_user_wallet = user['wallet']
                        print(f"new_user_wallet: {new_user_wallet}")
                        if new_user_wallet == previous_user_wallet:
                            user = get_user_details(user_id)
                            if user is None:
                                return None
                            previous_user_wallet = user['wallet']
                            print(f"previous wallet: {previous_user_wallet}")
                            new_balance = float(previous_user_wallet) - float(amount)
                            print(f"new_balance:{new_balance}")
                            doc_ref = user_collection.document(user_id)
                            doc_ref.update({'wallet': new_balance})
                            user = get_user_details(user_id)
                            new_user_wallet = user['wallet']
                            print(f"new_user_wallet: {new_user_wallet}")
                        else:
                            print("it's fine")

                    if user_details is not None:
                        print("yes")
                        first_name = user_details['first name']
                        print(first_name)
                        last_name = user_details['last name']
                        print(last_name)
                        email = user_details['email']
                        phone = user_details['phone']
                    else:
                        first_name = ""
                        last_name = ""
                        email = ""
                        phone = ""
                    details = {
                        'first_name': first_name,
                        'last_name': last_name,
                        'email': email,
                        'user_id': user_id
                    }

                    print("before url")

                    url = "https://cs.hubtel.com/commissionservices/2019604/fdd76c884e614b1c8f669a3207b09a98"
                    print(url)

                    print(f"receiver: {receiver}")

                    payload = json.dumps({
                        "Destination": receiver,
                        "Amount": amount,
                        "CallbackURL": "https://merchant.cloudhubgh.com/hubtel_webhook",
                        "ClientReference": reference
                    })
                    headers = {
                        'Authorization': config("HUBTEL_API_KEY"),
                        'Content-Type': 'application/json'
                    }

                    response = requests.request("POST", url, headers=headers, data=payload)

                    print(response.text)

                    if response.status_code == 200:
                        data = {
                            'batch_id': "unknown",
                            'buyer': phone,
                            'color_code': "Green",
                            'amount': amount,
                            'data_break_down': "",
                            'data_volume': "",
                            'date': date,
                            'date_and_time': date_and_time,
                            'done': "unknown",
                            'email': email,
                            'image': user_id,
                            'ishareBalance': 0,
                            'name': f"{first_name} {last_name}",
                            'number': receiver,
                            'paid_at': date_and_time,
                            'reference': reference,
                            'responseCode': "0",
                            'status': "Delivered",
                            'time': time,
                            'tranxId': str(tranx_id_generator()),
                            'type': "MTN Airtime",
                            'uid': user_id,
                            'bal': user_details["wallet"]
                        }
                        history_collection.document(date_and_time).set(data)
                        history_web.collection(email).document(date_and_time).set(data)
                        print("worked well")
                        return Response(
                            {"status": response.status_code, 'message': f'Transaction Completed Successfully'},
                        )
                    else:
                        print("not 200 error")
                        return Response({"status": response.status_code, 'message': f'Something went wrong'},
                                        status=status.HTTP_200_OK)
                else:
                    return Response({"status": 400, 'message': f'Insufficient Balance'},
                                    status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                print(e)
                return Response({"status": "400", 'message': f'Something went wrong: {e}'},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'error': 'Invalid Header Provided.'}, status=status.HTTP_401_UNAUTHORIZED)
    else:
        return Response({'error': 'Invalid Header Provided.'}, status=status.HTTP_401_UNAUTHORIZED)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@authentication_classes([BearerTokenAuthentication])
def initiate_voda_airtime(request):
    authorization_header = request.headers.get('Authorization')
    if authorization_header:
        auth_type, token = authorization_header.split(' ')
        if auth_type == 'Bearer':
            try:
                token_obj = Token.objects.get(key=token)
                user = token_obj.user
                user_id = user.user_id
                print(user_id)

                print("hiiiii")

                receiver = request.data.get('receiver')
                reference = request.data.get('reference')
                amount = request.data.get('amount')
                print(amount)
                # phone_number = request.data.get('phone_number')

                print("yo")

                if not receiver or not reference or not amount:
                    return Response({'message': 'Body parameters not valid. Check and try again.'},
                                    status=status.HTTP_400_BAD_REQUEST)

                print("got here")
                user_details = get_user_details(user_id)
                print(user_details['first name'])

                date = datetime.datetime.now().strftime("%a, %b %d, %Y")
                time = datetime.datetime.now().strftime("%I:%M:%S %p")
                date_and_time = datetime.datetime.now().isoformat()

                if "wallet" == "wallet":
                    try:
                        enough_balance = check_user_balance_against_price(user_id, amount)
                    except:
                        return Response(
                            {'code': '0001', 'message': f'User ID does not exist: User ID provided: {user_id}.'},
                            status=status.HTTP_400_BAD_REQUEST)
                else:
                    enough_balance = True
                    print("not wallet")
                print(enough_balance)
                if enough_balance:
                    user_details = get_user_details(user_id)
                    email = user_details['email']
                    print(enough_balance)
                    # hist = history_web.collection(email).document(date_and_time)
                    # doc = hist.get()
                    # if doc.exists:
                    #     return redirect(f"https://{callback_url}")
                    # else:
                    #     print("no record found")
                    if "wallet" == "wallet":
                        user = get_user_details(user_id)
                        if user is None:
                            return None
                        previous_user_wallet = user['wallet']
                        print(f"previous wallet: {previous_user_wallet}")
                        new_balance = float(previous_user_wallet) - float(amount)
                        print(f"new_balance:{new_balance}")
                        doc_ref = user_collection.document(user_id)
                        doc_ref.update({'wallet': new_balance})
                        user = get_user_details(user_id)
                        new_user_wallet = user['wallet']
                        print(f"new_user_wallet: {new_user_wallet}")
                        if new_user_wallet == previous_user_wallet:
                            user = get_user_details(user_id)
                            if user is None:
                                return None
                            previous_user_wallet = user['wallet']
                            print(f"previous wallet: {previous_user_wallet}")
                            new_balance = float(previous_user_wallet) - float(amount)
                            print(f"new_balance:{new_balance}")
                            doc_ref = user_collection.document(user_id)
                            doc_ref.update({'wallet': new_balance})
                            user = get_user_details(user_id)
                            new_user_wallet = user['wallet']
                            print(f"new_user_wallet: {new_user_wallet}")
                        else:
                            print("it's fine")

                if user_details is not None:
                    print("yes")
                    first_name = user_details['first name']
                    print(first_name)
                    last_name = user_details['last name']
                    print(last_name)
                    email = user_details['email']
                    phone = user_details['phone']
                else:
                    first_name = ""
                    last_name = ""
                    email = ""
                    phone = ""
                details = {
                    'first_name': first_name,
                    'last_name': last_name,
                    'email': email,
                    'user_id': user_id
                }

                url = "https://cs.hubtel.com/commissionservices/2019604/f4be83ad74c742e185224fdae1304800"

                payload = json.dumps({
                    "Destination": str(receiver),
                    "Amount": amount,
                    "CallbackURL": "https://merchant.cloudhubgh.com/hubtel_webhook",
                    "ClientReference": reference
                })
                headers = {
                    'Authorization': config("HUBTEL_API_KEY"),
                    'Content-Type': 'application/json'
                }

                response = requests.request("POST", url, headers=headers, data=payload)

                print(response.text)

                if response.status_code == 200:
                    data = {
                        'batch_id': "unknown",
                        'buyer': phone,
                        'color_code': "Green",
                        'amount': amount,
                        'data_break_down': "",
                        'data_volume': "",
                        'date': date,
                        'date_and_time': date_and_time,
                        'done': "unknown",
                        'email': email,
                        'image': user_id,
                        'ishareBalance': 0,
                        'name': f"{first_name} {last_name}",
                        'number': receiver,
                        'paid_at': date_and_time,
                        'reference': reference,
                        'responseCode': "0",
                        'status': "Delivered",
                        'time': time,
                        'tranxId': str(tranx_id_generator()),
                        'type': "Vodafone Airtime",
                        'uid': user_id,
                        'bal': user["wallet"]
                    }
                    history_collection.document(date_and_time).set(data)
                    history_web.collection(email).document(date_and_time).set(data)
                    print("worked well")
                    return Response({"status": response.status_code, 'message': f'Transaction Completed Successfully'},
                                    )
                else:
                    print("not 200 error")
                    return Response({"status": response.status_code, 'message': f'Something went wrong'},
                                    status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"status": '400', 'message': f'Something went wrong: {e}'},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'error': 'Invalid Header Provided.'}, status=status.HTTP_401_UNAUTHORIZED)
    else:
        return Response({'error': 'Invalid Header Provided.'}, status=status.HTTP_401_UNAUTHORIZED)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@authentication_classes([BearerTokenAuthentication])
def admin_initiate_voda_airtime(request):
    authorization_header = request.headers.get('Authorization')
    if authorization_header:
        auth_type, token = authorization_header.split(' ')
        if auth_type == 'Bearer':
            try:
                token_obj = Token.objects.get(key=token)
                token_key = token_obj.key
                if token_key != config("TOKEN_KEY"):
                    return Response({'message': 'Authorisation Failed.'},
                                    status=status.HTTP_400_BAD_REQUEST)

                receiver = request.data.get('receiver')
                reference = request.data.get('reference')
                amount = request.data.get('amount')
                user_id = request.data.get('user_id')
                print(amount)
                # phone_number = request.data.get('phone_number')

                print("yo")

                if not receiver or not reference or not amount:
                    return Response({'message': 'Body parameters not valid. Check and try again.'},
                                    status=status.HTTP_400_BAD_REQUEST)

                print("got here")
                user_details = get_user_details(user_id)
                print(user_details['first name'])

                date = datetime.datetime.now().strftime("%a, %b %d, %Y")
                time = datetime.datetime.now().strftime("%I:%M:%S %p")
                date_and_time = datetime.datetime.now().isoformat()

                if "wallet" == "wallet":
                    try:
                        enough_balance = check_user_balance_against_price(user_id, amount)
                    except:
                        return Response(
                            {'code': '0001', 'message': f'User ID does not exist: User ID provided: {user_id}.'},
                            status=status.HTTP_400_BAD_REQUEST)
                else:
                    enough_balance = True
                    print("not wallet")
                print(enough_balance)
                if enough_balance:
                    user_details = get_user_details(user_id)
                    email = user_details['email']
                    print(enough_balance)
                    # hist = history_web.collection(email).document(date_and_time)
                    # doc = hist.get()
                    # if doc.exists:
                    #     return redirect(f"https://{callback_url}")
                    # else:
                    #     print("no record found")
                    if "wallet" == "wallet":
                        user = get_user_details(user_id)
                        if user is None:
                            return None
                        previous_user_wallet = user['wallet']
                        print(f"previous wallet: {previous_user_wallet}")
                        new_balance = float(previous_user_wallet) - float(amount)
                        print(f"new_balance:{new_balance}")
                        doc_ref = user_collection.document(user_id)
                        doc_ref.update({'wallet': new_balance})
                        user = get_user_details(user_id)
                        new_user_wallet = user['wallet']
                        print(f"new_user_wallet: {new_user_wallet}")
                        if new_user_wallet == previous_user_wallet:
                            user = get_user_details(user_id)
                            if user is None:
                                return None
                            previous_user_wallet = user['wallet']
                            print(f"previous wallet: {previous_user_wallet}")
                            new_balance = float(previous_user_wallet) - float(amount)
                            print(f"new_balance:{new_balance}")
                            doc_ref = user_collection.document(user_id)
                            doc_ref.update({'wallet': new_balance})
                            user = get_user_details(user_id)
                            new_user_wallet = user['wallet']
                            print(f"new_user_wallet: {new_user_wallet}")
                        else:
                            print("it's fine")

                    if user_details is not None:
                        print("yes")
                        first_name = user_details['first name']
                        print(first_name)
                        last_name = user_details['last name']
                        print(last_name)
                        email = user_details['email']
                        phone = user_details['phone']
                    else:
                        first_name = ""
                        last_name = ""
                        email = ""
                        phone = ""
                    details = {
                        'first_name': first_name,
                        'last_name': last_name,
                        'email': email,
                        'user_id': user_id
                    }

                    url = "https://cs.hubtel.com/commissionservices/2019604/f4be83ad74c742e185224fdae1304800"

                    payload = json.dumps({
                        "Destination": str(receiver),
                        "Amount": amount,
                        "CallbackURL": "https://merchant.cloudhubgh.com/hubtel_webhook",
                        "ClientReference": reference
                    })
                    headers = {
                        'Authorization': config("HUBTEL_API_KEY"),
                        'Content-Type': 'application/json'
                    }

                    response = requests.request("POST", url, headers=headers, data=payload)

                    print(response.text)

                    if response.status_code == 200:
                        data = {
                            'batch_id': "unknown",
                            'buyer': phone,
                            'color_code': "Green",
                            'amount': amount,
                            'data_break_down': "",
                            'data_volume': "",
                            'date': date,
                            'date_and_time': date_and_time,
                            'done': "unknown",
                            'email': email,
                            'image': user_id,
                            'ishareBalance': 0,
                            'name': f"{first_name} {last_name}",
                            'number': receiver,
                            'paid_at': date_and_time,
                            'reference': reference,
                            'responseCode': "0",
                            'status': "Delivered",
                            'time': time,
                            'tranxId': str(tranx_id_generator()),
                            'type': "Vodafone Airtime",
                            'uid': user_id,
                            'bal': user_details["wallet"]
                        }
                        history_collection.document(date_and_time).set(data)
                        history_web.collection(email).document(date_and_time).set(data)
                        print("worked well")
                        return Response(
                            {"status": response.status_code, 'message': f'Transaction Completed Successfully'},
                        )
                    else:
                        print("not 200 error")
                        return Response({"status": response.status_code, 'message': f'Something went wrong'},
                                        status=status.HTTP_200_OK)
                else:
                    return Response({"status": 400, 'message': f'Insufficient Balance'},
                                    status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                print(e)
                return Response({"status": '400', 'message': f'Something went wrong: {e}'},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'error': 'Invalid Header Provided.'}, status=status.HTTP_401_UNAUTHORIZED)
    else:
        return Response({'error': 'Invalid Header Provided.'}, status=status.HTTP_401_UNAUTHORIZED)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@authentication_classes([BearerTokenAuthentication])
def initiate_glo_airtime(request):
    authorization_header = request.headers.get('Authorization')
    if authorization_header:
        auth_type, token = authorization_header.split(' ')
        if auth_type == 'Bearer':
            try:
                token_obj = Token.objects.get(key=token)
                user = token_obj.user
                user_id = user.user_id
                print(user_id)

                print("hiiiii")

                receiver = request.data.get('receiver')
                reference = request.data.get('reference')
                amount = request.data.get('amount')
                print(amount)
                # phone_number = request.data.get('phone_number')

                print("yo")

                if not receiver or not reference or not amount:
                    return Response({'message': 'Body parameters not valid. Check and try again.'},
                                    status=status.HTTP_400_BAD_REQUEST)

                print("got here")
                user_details = get_user_details(user_id)
                print(user_details['first name'])

                date = datetime.datetime.now().strftime("%a, %b %d, %Y")
                time = datetime.datetime.now().strftime("%I:%M:%S %p")
                date_and_time = datetime.datetime.now().isoformat()

                if "wallet" == "wallet":
                    try:
                        enough_balance = check_user_balance_against_price(user_id, amount)
                    except:
                        return Response(
                            {'code': '0001', 'message': f'User ID does not exist: User ID provided: {user_id}.'},
                            status=status.HTTP_400_BAD_REQUEST)
                else:
                    enough_balance = True
                    print("not wallet")
                print(enough_balance)
                if enough_balance:
                    user_details = get_user_details(user_id)
                    email = user_details['email']
                    print(enough_balance)
                    # hist = history_web.collection(email).document(date_and_time)
                    # doc = hist.get()
                    # if doc.exists:
                    #     return redirect(f"https://{callback_url}")
                    # else:
                    #     print("no record found")
                    if "wallet" == "wallet":
                        user = get_user_details(user_id)
                        if user is None:
                            return None
                        previous_user_wallet = user['wallet']
                        print(f"previous wallet: {previous_user_wallet}")
                        new_balance = float(previous_user_wallet) - float(amount)
                        print(f"new_balance:{new_balance}")
                        doc_ref = user_collection.document(user_id)
                        doc_ref.update({'wallet': new_balance})
                        user = get_user_details(user_id)
                        new_user_wallet = user['wallet']
                        print(f"new_user_wallet: {new_user_wallet}")
                        if new_user_wallet == previous_user_wallet:
                            user = get_user_details(user_id)
                            if user is None:
                                return None
                            previous_user_wallet = user['wallet']
                            print(f"previous wallet: {previous_user_wallet}")
                            new_balance = float(previous_user_wallet) - float(amount)
                            print(f"new_balance:{new_balance}")
                            doc_ref = user_collection.document(user_id)
                            doc_ref.update({'wallet': new_balance})
                            user = get_user_details(user_id)
                            new_user_wallet = user['wallet']
                            print(f"new_user_wallet: {new_user_wallet}")
                        else:
                            print("it's fine")

                if user_details is not None:
                    print("yes")
                    first_name = user_details['first name']
                    print(first_name)
                    last_name = user_details['last name']
                    print(last_name)
                    email = user_details['email']
                    phone = user_details['phone']
                else:
                    first_name = ""
                    last_name = ""
                    email = ""
                    phone = ""
                details = {
                    'first_name': first_name,
                    'last_name': last_name,
                    'email': email,
                    'user_id': user_id
                }

                url = "https://cs.hubtel.com/commissionservices/2019604/47d88e88f50f47468a34a14ac73e8ab5"

                payload = json.dumps({
                    "Destination": str(receiver),
                    "Amount": amount,
                    "CallbackURL": "https://merchant.cloudhubgh.com/hubtel_webhook",
                    "ClientReference": reference
                })
                headers = {
                    'Authorization': config("HUBTEL_API_KEY"),
                    'Content-Type': 'application/json'
                }

                response = requests.request("POST", url, headers=headers, data=payload)

                print(response.text)

                if response.status_code == 200:
                    data = {
                        'batch_id': "unknown",
                        'buyer': phone,
                        'color_code': "Green",
                        'amount': amount,
                        'data_break_down': "",
                        'data_volume': "",
                        'date': date,
                        'date_and_time': date_and_time,
                        'done': "unknown",
                        'email': email,
                        'image': user_id,
                        'ishareBalance': 0,
                        'name': f"{first_name} {last_name}",
                        'number': receiver,
                        'paid_at': date_and_time,
                        'reference': reference,
                        'responseCode': "0",
                        'status': "Delivered",
                        'time': time,
                        'tranxId': str(tranx_id_generator()),
                        'type': "Glo GH Airtime",
                        'uid': user_id,
                        'bal': user["wallet"]
                    }
                    history_collection.document(date_and_time).set(data)
                    history_web.collection(email).document(date_and_time).set(data)
                    print("worked well")
                    return Response({"status": response.status_code, 'message': f'Transaction Completed Successfully'},
                                    )
                else:
                    print("not 200 error")
                    return Response({"status": response.status_code, 'message': f'Something went wrong'},
                                    status=status.HTTP_200_OK)
            except Exception as e:
                print(e)
                return Response({"status": '400', 'message': f'Something went wrong: {e}'},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'error': 'Invalid Header Provided.'}, status=status.HTTP_401_UNAUTHORIZED)
    else:
        return Response({'error': 'Invalid Header Provided.'}, status=status.HTTP_401_UNAUTHORIZED)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@authentication_classes([BearerTokenAuthentication])
def admin_initiate_glo_airtime(request):
    authorization_header = request.headers.get('Authorization')
    if authorization_header:
        auth_type, token = authorization_header.split(' ')
        if auth_type == 'Bearer':
            try:
                token_obj = Token.objects.get(key=token)
                token_key = token_obj.key
                if token_key != config("TOKEN_KEY"):
                    return Response({'message': 'Authorisation Failed.'},
                                    status=status.HTTP_400_BAD_REQUEST)

                receiver = request.data.get('receiver')
                reference = request.data.get('reference')
                amount = request.data.get('amount')
                user_id = request.data.get('user_id')
                print(amount)
                # phone_number = request.data.get('phone_number')

                print("yo")

                if not receiver or not reference or not amount:
                    return Response({'message': 'Body parameters not valid. Check and try again.'},
                                    status=status.HTTP_400_BAD_REQUEST)

                print("got here")
                user_details = get_user_details(user_id)
                print(user_details['first name'])

                date = datetime.datetime.now().strftime("%a, %b %d, %Y")
                time = datetime.datetime.now().strftime("%I:%M:%S %p")
                date_and_time = datetime.datetime.now().isoformat()

                if "wallet" == "wallet":
                    try:
                        enough_balance = check_user_balance_against_price(user_id, amount)
                    except:
                        return Response(
                            {'code': '0001', 'message': f'User ID does not exist: User ID provided: {user_id}.'},
                            status=status.HTTP_400_BAD_REQUEST)
                else:
                    enough_balance = True
                    print("not wallet")
                print(enough_balance)
                if enough_balance:
                    user_details = get_user_details(user_id)
                    email = user_details['email']
                    print(enough_balance)
                    # hist = history_web.collection(email).document(date_and_time)
                    # doc = hist.get()
                    # if doc.exists:
                    #     return redirect(f"https://{callback_url}")
                    # else:
                    #     print("no record found")
                    if "wallet" == "wallet":
                        user = get_user_details(user_id)
                        if user is None:
                            return None
                        previous_user_wallet = user['wallet']
                        print(f"previous wallet: {previous_user_wallet}")
                        new_balance = float(previous_user_wallet) - float(amount)
                        print(f"new_balance:{new_balance}")
                        doc_ref = user_collection.document(user_id)
                        doc_ref.update({'wallet': new_balance})
                        user = get_user_details(user_id)
                        new_user_wallet = user['wallet']
                        print(f"new_user_wallet: {new_user_wallet}")
                        if new_user_wallet == previous_user_wallet:
                            user = get_user_details(user_id)
                            if user is None:
                                return None
                            previous_user_wallet = user['wallet']
                            print(f"previous wallet: {previous_user_wallet}")
                            new_balance = float(previous_user_wallet) - float(amount)
                            print(f"new_balance:{new_balance}")
                            doc_ref = user_collection.document(user_id)
                            doc_ref.update({'wallet': new_balance})
                            user = get_user_details(user_id)
                            new_user_wallet = user['wallet']
                            print(f"new_user_wallet: {new_user_wallet}")
                        else:
                            print("it's fine")

                        if user_details is not None:
                            print("yes")
                            first_name = user_details['first name']
                            print(first_name)
                            last_name = user_details['last name']
                            print(last_name)
                            email = user_details['email']
                            phone = user_details['phone']
                        else:
                            first_name = ""
                            last_name = ""
                            email = ""
                            phone = ""
                        details = {
                            'first_name': first_name,
                            'last_name': last_name,
                            'email': email,
                            'user_id': user_id
                        }

                        url = "https://cs.hubtel.com/commissionservices/2019604/47d88e88f50f47468a34a14ac73e8ab5"

                        payload = json.dumps({
                            "Destination": str(receiver),
                            "Amount": amount,
                            "CallbackURL": "https://merchant.cloudhubgh.com/hubtel_webhook",
                            "ClientReference": reference
                        })
                        headers = {
                            'Authorization': config("HUBTEL_API_KEY"),
                            'Content-Type': 'application/json'
                        }

                        response = requests.request("POST", url, headers=headers, data=payload)

                        print(response.text)

                        if response.status_code == 200:
                            data = {
                                'batch_id': "unknown",
                                'buyer': phone,
                                'color_code': "Green",
                                'amount': amount,
                                'data_break_down': "",
                                'data_volume': "",
                                'date': date,
                                'date_and_time': date_and_time,
                                'done': "unknown",
                                'email': email,
                                'image': user_id,
                                'ishareBalance': 0,
                                'name': f"{first_name} {last_name}",
                                'number': receiver,
                                'paid_at': date_and_time,
                                'reference': reference,
                                'responseCode': "0",
                                'status': "Delivered",
                                'time': time,
                                'tranxId': str(tranx_id_generator()),
                                'type': "Glo GH Airtime",
                                'uid': user_id,
                                'bal': user_details["wallet"]
                            }
                            history_collection.document(date_and_time).set(data)
                            history_web.collection(email).document(date_and_time).set(data)
                            print("worked well")
                            return Response(
                                {"status": response.status_code, 'message': f'Transaction Completed Successfully'},
                            )
                        else:
                            print("not 200 error")
                            return Response({"status": response.status_code, 'message': f'Something went wrong'},
                                            status=status.HTTP_200_OK)
                    else:
                        return Response({"status": 400, 'message': f'Insufficient Balance'},
                                        status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                print(e)
                return Response({"status": '400', 'message': f'Something went wrong: {e}'},
                                status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'error': 'Invalid Header Provided.'}, status=status.HTTP_401_UNAUTHORIZED)
    else:
        return Response({'error': 'Invalid Header Provided.'}, status=status.HTTP_401_UNAUTHORIZED)


from django.http import HttpResponse
from io import BytesIO
from .models import MTNTransaction  # Adjust the import based on your model's location

from openpyxl import load_workbook


@csrf_exempt
def export_unknown_transactions(request):
    existing_excel_path = 'business_api/ALL PACKAGES LATEST.xlsx'  # Update with your file path

    # Load the existing Excel file using openpyxl.Workbook
    book = load_workbook(existing_excel_path)

    # Get the active sheet
    sheet_name = 'Sheet1'
    sheet = book[sheet_name] if sheet_name in book.sheetnames else book.active

    # Clear existing data from the sheet (excluding headers)
    for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):
        for cell in row:
            cell.value = None

    # Query your Django model for the first 200 records with batch_id 'Unknown' and ordered by status and date
    queryset = MTNTransaction.objects.filter(batch_id='Unknown', status="Undelivered")[:50]

    # Process transactions with batch_id 'Unknown'
    counter = 0

    for record in queryset:
        print(counter)

        # Extract required fields from your Django model
        bundle_volume_mb = record.bundle_volume  # Assuming a default of 0 if datavolume is missing
        number = str(record.number)  # Convert to string to keep leading zeros

        # Convert datavolume from MB to GB
        bundle_volume_gb = round(float(bundle_volume_mb) / 1000)

        # Find the row index where you want to populate the data (adjust as needed)
        target_row = 2 + counter  # Assuming the data starts from row 2

        # Populate the specific cells with the new data
        sheet.cell(row=target_row, column=1, value=number)  # Keep leading zeros
        sheet.cell(row=target_row, column=2, value=float(bundle_volume_gb))  # Convert to float

        # Update 'batch_id' to 'processing' in your Django model
        record.batch_id = 'accepted'
        record.status = 'Processing'
        record.save()

        counter += 1
        txn = mtn_other.document(record.firebase_date)
        txn.update({'batch_id': 'accepted', 'status': 'Processing'})

    print(f"Total transactions to export: {counter}")

    # Save changes to the existing Excel file
    book.save(existing_excel_path)

    # You can continue with the response as needed
    excel_buffer = BytesIO()

    # Save the workbook to the buffer
    book.save(excel_buffer)

    # Create a response with the Excel file
    response = HttpResponse(excel_buffer.getvalue(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={datetime.datetime.now()}.xlsx'

    return response
